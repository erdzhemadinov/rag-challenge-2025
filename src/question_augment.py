# python
"""
Minimal question augmentation module.

Usage:
    from question_augment import (
        GPTQuestionAugmentor, augment_questions_xlsx, probe_augmentor
    )
"""

from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple, Optional
import os
import sys

try:
    from openai import OpenAI
except ImportError:  # defer hard failure until actually used
    OpenAI = None  # type: ignore

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

_AUG_SYSTEM_MSG = (
    "Ты модуль минимального уточнения финансовых вопросов.\n"
    "Добавь при необходимости одно краткое количественное уточнение: 'всего', "
    "'общее количество', 'общая', 'совокупная', если оно логически подразумевается и отсутствует.\n"
    "Правила:\n"
    "- Сохраняй исходную пунктуацию и порядок слов, кроме точечной вставки.\n"
    "- Не дублируй уже имеющиеся слова: всего, итого, общее, общая, общий, совокупная.\n"
    "- Собственные имена, аббревиатуры и знаки оставляй без изменений.\n"
    "- Не перефразируй полностью.\n"
    "- Если вставка не нужна или риск искажения смысла — верни оригинал.\n"
    "- Ответ: одна строка (модифицированный или оригинальный вопрос)."
)

_ALREADY_Q = {"всего", "итого", "общее", "общая", "общий", "совокуп", "aggregate", "total"}


@dataclass
class GPTQuestionAugmentor:
    """
    Lightweight wrapper for minimal augmentation of Russian financial questions.
    """
    model: str = "gpt-4o"
    temperature: float = 0.0
    client: Optional[OpenAI] = None
    verbose: bool = False

    def __post_init__(self) -> None:
        if self.client is None:
            if OpenAI is None:
                raise RuntimeError("Install dependency: pip install openai")
            self.client = OpenAI()
        self.model = (self.model or "gpt-4o").strip()

    # --- heuristics ---
    def needs_augmentation(self, q: str) -> bool:
        if not q or not q.strip():
            return False
        low = q.lower()
        if any(tok in low for tok in _ALREADY_Q):
            return False
        return low.startswith((
            "сколько", "какова", "каков", "число",
            "количество", "на какую дату"
        ))

    def _validate(self, original: str, augmented: str) -> str:
        aug = (augmented or "").strip()
        if not aug:
            return original
        # preserve terminal punctuation
        for ch in "?!.…":
            if original.rstrip().endswith(ch) and not aug.rstrip().endswith(ch):
                return original
        # length guard
        if len(aug) > len(original) + 25:
            return original
        return aug

    # --- main API ---
    def augment_one(self, question: str) -> str:
        if not self.needs_augmentation(question):
            if self.verbose:
                print(f"[augment][skip] {question}")
            return question
        try:
            resp = self.client.chat.completions.create(
                model=self.model,
                #temperature=self.temperature,
                messages=[
                    {"role": "system", "content": _AUG_SYSTEM_MSG},
                    {"role": "user", "content": question},
                ],
            )
            candidate = resp.choices[0].message.content
        except Exception as e:
            print(f"[augment][error] model={self.model} question='{question}': {e}", file=sys.stderr)
            return question
        out = self._validate(question, candidate)
        if self.verbose:
            if out != question:
                print(f"[augment][ok] '{question}' -> '{out}'")
            else:
                print(f"[augment][nochange] '{question}'")
        return out


def augment_questions_xlsx(
    src: Path,
    dst: Path,
    augmentor: GPTQuestionAugmentor,
    question_header_substrings: Tuple[str, ...] = ("question", "вопрос"),
) -> Tuple[Path, List[Tuple[int, str, str]]]:
    """
    Read XLSX at `src`, augment qualifying questions in-place into a copy at `dst`.
    Returns (dst_path, changes) where changes is a list of (row_number, original, augmented).
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl required: pip install openpyxl")

    wb = openpyxl.load_workbook(src)
    ws = wb.active

    # detect question column
    first_row_vals = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = 0
    header_map = {}
    for idx, val in enumerate(first_row_vals):
        if val is None:
            continue
        norm = str(val).strip().lower()
        header_map[norm] = idx
    for h, idx in header_map.items():
        if any(sub in h for sub in question_header_substrings):
            header_index = idx
            break

    changes: List[Tuple[int, str, str]] = []
    for row in ws.iter_rows(min_row=2):
        cell = row[header_index]
        if isinstance(cell.value, str) and cell.value.strip():
            original = cell.value
            augmented = augmentor.augment_one(original)
            cell.value = augmented
            if augmented != original:
                changes.append((cell.row, original, augmented))

    wb.save(dst)

    if augmentor.verbose:
        if changes:
            print("[augment] Processed questions:")
            for r, o, a in changes:
                print(f"  Row {r}: {o} -> {a}")
        else:
            print("[augment] No questions required augmentation.")
    return dst, changes


def probe_augmentor(augmentor: GPTQuestionAugmentor, sample: str = "Сколько акций?") -> bool:
    """
    Simple probe to verify the model responds. Returns True on success.
    """
    try:
        _ = augmentor.augment_one(sample)
        return True
    except Exception as e:
        print(f"[augment][probe][fail] {e}", file=sys.stderr)
        return False


__all__ = [
    "GPTQuestionAugmentor",
    "augment_questions_xlsx",
    "probe_augmentor",
]