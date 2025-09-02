
"""
pipeline
"""

from __future__ import annotations

import os
import pickle
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
from dotenv import load_dotenv
from tqdm import tqdm
load_dotenv()

OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]
# logs
log = logging.getLogger("RAGPipeline")
if not log.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")

# libs
try:
    import faiss  # type: ignore
except Exception as e:
    raise RuntimeError("faiss-cpu не установлен. Установите: pip install faiss-cpu") from e

try:
    from openai import OpenAI
except Exception as e:
    raise RuntimeError("openai>=1.40.0 не установлен. Установите: pip install openai>=1.40.0") from e

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import normalize as sk_normalize


# ==========================
# Configs
# ==========================

@dataclass
class CleanRules:
    normalize_spaces: bool = True
    strip_control_chars: bool = True

@dataclass
class ParseConfig:
    input_dir: Path = Path("input")
    output_dir: Path = Path("parsed")
    docling_backend: bool = True               # попытаться использовать pdf_parsing.PDFParser
    csv_metadata_path: Optional[Path] = None
    debug_json_dump_dir: Optional[Path] = None
    page_images_dpi: int = 300                 # для fallback
    recurse: bool = True

@dataclass
class ChunkingConfig:
    # fixed режим
    chunk_size_tokens: int = 300
    chunk_overlap_tokens: int = 50
    # auto режим
    mode: str = "fixed"               # "fixed" | "auto"
    min_chunk_tokens: int = 300
    max_chunk_tokens: int = 800
    overlap_ratio: float = 0.2        # для auto

@dataclass
class EmbeddingConfig:
    openai_api_key_env: str = "OPENAI_API_KEY"
    model: str = "text-embedding-3-large"
    batch_size: int = 64

@dataclass
class HybridConfig:
    enable: bool = True
    weight_faiss: float = 0.7
    weight_lexical: float = 0.3
    max_features: int = 200_000

@dataclass
class IndexerConfig:
    parsed_dir: Path = Path("parsed")
    index_dir: Path = Path("indices_cs300_ov50")
    chunking: ChunkingConfig = field(default_factory=ChunkingConfig)
    embed: EmbeddingConfig = field(default_factory=EmbeddingConfig)
    hybrid: HybridConfig = field(default_factory=HybridConfig)

@dataclass
class RetrieverConfig:
    index_dir: Path = Path("indices_cs300_ov50")
    embed: EmbeddingConfig = field(default_factory=EmbeddingConfig)
    unique_by_page: bool = True
    top_k: int = 10
    hybrid: HybridConfig = field(default_factory=HybridConfig)

@dataclass
class RerankConfig:
    enable: bool = True
    top_k_pre_rerank: int = 30
    llm_model: str = "gpt-4o-mini"
    openai_api_key_env: str = "OPENAI_API_KEY"
    weight_faiss: float = 0.3
    weight_llm: float = 0.7
    batch_size: int = 10
    temperature: float = 0.0
    seed: int = 1015
    n: int = 3
@dataclass
class QuestionPrepConfig:
    new_challenge_pipeline: bool = False
    subset_csv_path: Optional[Path] = None
    enable_comparatives: bool = True


# ==========================
# Cleaning helpers
# ==========================

def _normalize_ws(text: str) -> str:
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\u00A0", " ", text)
    text = re.sub(r"\s+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

def _strip_control_chars(text: str) -> str:
    return "".join(ch for ch in text if ch == "\n" or ch >= " ")

def clean_text(text: str, rules: CleanRules) -> str:
    if not text:
        return ""
    if rules.strip_control_chars:
        text = _strip_control_chars(text)
    if rules.normalize_spaces:
        text = _normalize_ws(text)
    return text


# ==========================
# Token utilities
# ==========================

def _approx_tokens_count(s: str) -> int:
    if not s:
        return 0
    return max(1, (len(s) + 3) // 4)  # ~4 chars/token

def _split_by_tokens(text: str, size: int, overlap: int):
    if not text.strip():
        return []
    chars_per_tok = 4
    step_c = size * chars_per_tok
    ov_c = overlap * chars_per_tok
    out = []
    i = 0
    N = len(text)
    while i < N:
        j = min(N, i + step_c)
        out.append((i, j, text[i:j]))
        if j == N:
            break
        i = max(0, j - ov_c)
    return out

def _estimate_auto_chunk_size(text: str, cfg: ChunkingConfig) -> Tuple[int, int]:
    nl = text.count("\n")
    pipes = text.count("|")
    tabs = text.count("\t")
    bullets = sum(text.count(b) for b in ["•", "- ", "– ", "* "])
    lines = [len(l.strip()) for l in text.splitlines() if l.strip()]
    avg_line = (sum(lines) / max(1, len(lines))) if lines else 0.0
    structure_score = (nl * 0.5 + pipes * 3 + tabs * 2 + bullets * 1.5) / max(1, len(text) / 1000.0)
    lo, hi = cfg.min_chunk_tokens, cfg.max_chunk_tokens
    if structure_score >= 6:
        size = hi
    elif structure_score >= 3:
        size = int((lo + hi) / 2)
    else:
        size = lo
    if avg_line <= 40:
        size = min(hi, int(size * 1.2))
    elif avg_line >= 140:
        size = max(lo, int(size * 0.85))
    overlap = max(0, int(size * float(cfg.overlap_ratio)))
    return size, overlap

def split_into_chunks_smart(text: str, cfg: ChunkingConfig):
    if (cfg.mode or "fixed").lower() != "auto":
        return _split_by_tokens(text, cfg.chunk_size_tokens, cfg.chunk_overlap_tokens)
    size, overlap = _estimate_auto_chunk_size(text, cfg)
    return _split_by_tokens(text, size, overlap)


# 2) Keep this helper once (place it above CorpusParser or anywhere before use)
def _hybrid_ocr_postprocess_pages(
    docling_pages: List[Dict[str, Any]],
    pdf_path: Path,
    *,
    min_chars: int = 200,
    min_confidence: float = 0.60,
    openai_model: str = "gpt-4o-mini",
    dpi: int = 300,
    jpeg_quality: int = 80,
) -> List[Dict[str, Any]]:
    """
    Replace page text with OpenAI OCR if Docling text is too short/low-confidence.
    Expects each page dict to have 'text' and optionally 'confidence' (or 'ocr_confidence').
    """
    for i, p in enumerate(docling_pages):
        dl_text = (p.get("text") or "")
        dl_conf = p.get("confidence", p.get("ocr_confidence"))
        new_text, source = combine_docling_and_openai_for_page(
            str(pdf_path),
            i,
            dl_text,
            dl_conf,
            min_chars=min_chars,
            min_confidence=min_confidence,
            openai_model=openai_model,
            dpi=dpi,
            jpeg_quality=jpeg_quality,
        )
        p["text"] = new_text
        p["ocr_source"] = source
    return docling_pages


# 3) Inside CorpusParser.run(), after Docling export, call the hybrid postprocess
class CorpusParser:
    def __init__(self, cfg: ParseConfig):
        self.cfg = cfg
        self.parsed_dir = Path(cfg.output_dir)
        self.parsed_dir.mkdir(parents=True, exist_ok=True)
        self._pdf_parser = None
        if cfg.docling_backend:
            try:
                from pdf_parsing import PDFParser as DLParser  # type: ignore
                self._pdf_parser = DLParser(
                    output_dir=self.parsed_dir,
                    csv_metadata_path=cfg.csv_metadata_path
                )
                if cfg.debug_json_dump_dir:
                    self._pdf_parser.debug_data_path = Path(cfg.debug_json_dump_dir)
            except Exception as e:
                log.warning("Docling-бэкенд недоступен: %s. Использую fallback на PyMuPDF.", e)
                self._pdf_parser = None

    def run(self):
        input_dir = Path(self.cfg.input_dir)
        pdfs = list(input_dir.rglob("*.pdf")) if self.cfg.recurse else list(input_dir.glob("*.pdf"))
        if not pdfs:
            log.warning("PDF не найдены в %s", input_dir)
            return []
        if self._pdf_parser is not None:
            log.info("Парсинг через pdf_parsing.PDFParser: %d файлов", len(pdfs))
            self._pdf_parser.parse_and_export(input_doc_paths=pdfs)

            # NEW: Hybrid OCR post-process Docling JSONs per original PDF
            for src_pdf in tqdm(pdfs, desc="Hybrid OCR postprocess"):
                json_path = self.parsed_dir / f"{src_pdf.stem}.json"
                if not json_path.exists():
                    continue
                try:
                    data = json.loads(json_path.read_text(encoding="utf-8"))
                    # Build page list with aggregated text
                    pages_in = []
                    for p in data.get("content", []):
                        buf = []
                        for item in p.get("content", []):
                            if item.get("type") == "text":
                                buf.append(item.get("text", ""))
                        pages_in.append({"text": "\n".join(buf)})
                    # Apply hybrid OCR selection
                    pages_out = _hybrid_ocr_postprocess_pages(pages_in, src_pdf)
                    # Write back as a single text item per page + mark source
                    for i, p in enumerate(data.get("content", [])):
                        new_text = pages_out[i].get("text", "")
                        ocr_src = pages_out[i].get("ocr_source", "docling")
                        p["content"] = [{"type": "text", "text": new_text}]
                        p["ocr_source"] = ocr_src
                    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
                except Exception as e:
                    log.warning("Hybrid OCR postprocess failed for %s: %s", json_path.name, e)
        else:
            import fitz  # PyMuPDF
            for p in tqdm(pdfs, desc="Fallback: PyMuPDF parse (no OCR)"):
                try:
                    doc = fitz.open(p)
                    content = []
                    for i in range(len(doc)):
                        page = doc.load_page(i)
                        text = page.get_text()
                        content.append({
                            "page": i+1,
                            "content": [{"type": "text", "text": text}],
                            "page_dimensions": {}
                        })
                    out = {
                        "metainfo": {"sha1_name": p.stem, "pages_amount": len(doc)},
                        "content": content,
                        "tables": [],
                        "pictures": []
                    }
                    with (self.parsed_dir / f"{p.stem}.json").open("w", encoding="utf-8") as f:
                        json.dump(out, f, ensure_ascii=False, indent=2)
                except Exception as e:
                    log.error("Ошибка парсинга %s: %s", p, e)
        return list(self.parsed_dir.glob("*.json"))

# ==========================
# Indexer
# ==========================

class RAGIndexer:
    def __init__(self, cfg: IndexerConfig):
        self.cfg = cfg
        self.client = OpenAI()
        self.index_dir = Path(cfg.index_dir)
        self.index_dir.mkdir(parents=True, exist_ok=True)

    def _load_pages_from_parsed(self, report_json: Dict[str, Any]) -> List[Dict[str, Any]]:
        pages = []
        for p in report_json.get("content", []):
            page_num = p.get("page", None)
            buf = []
            for item in p.get("content", []):
                if item.get("type") == "text":
                    buf.append(item.get("text", ""))
            page_text = "\n".join(buf)
            pages.append({"page": page_num, "text": page_text})
        return pages

    def _dedup_chunks(self, chunks):
        seen = set()
        unique = []
        for st, en, t in chunks:
            key = hash(t.strip()[:256])
            if key in seen:
                continue
            seen.add(key)
            unique.append((st, en, t))
        return unique

    def _embed_texts(self, texts: List[str]) -> np.ndarray:
        em_cfg = self.cfg.embed
        api_key = os.getenv(em_cfg.openai_api_key_env)
        if not api_key:
            raise RuntimeError(f"{em_cfg.openai_api_key_env} is not set")
        out_vecs: List[List[float]] = []
        for i in range(0, len(texts), em_cfg.batch_size):
            batch = texts[i:i+em_cfg.batch_size]
            resp = self.client.embeddings.create(model=em_cfg.model, input=batch)
            for d in resp.data:
                out_vecs.append(d.embedding)
        arr = np.array(out_vecs, dtype="float32")
        norms = np.linalg.norm(arr, axis=1, keepdims=True) + 1e-12
        arr = arr / norms
        return arr

    def _build_doc_index(self, json_path: Path) -> Optional[Path]:
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception as e:
            log.error("Не удалось прочитать %s: %s", json_path, e)
            return None

        pages = self._load_pages_from_parsed(data)
        if not pages:
            log.warning("Нет контента в %s", json_path.name)
            return None

        clean_rules = CleanRules()
        chunks_texts: List[str] = []
        chunks_meta: List[Dict[str, Any]] = []
        for p in pages:
            text = clean_text(p["text"], clean_rules)
            chs = split_into_chunks_smart(text, self.cfg.chunking)
            chs = self._dedup_chunks(chs)
            for st, en, ch in chs:
                if not ch.strip():
                    continue
                chunks_texts.append(ch)
                chunks_meta.append({"page_id": p["page"], "start": st, "end": en, "text": ch})

        if not chunks_texts:
            log.warning("Пустые чанки в %s", json_path.name)
            return None

        emb = self._embed_texts(chunks_texts)

        dim = emb.shape[1]
        index = faiss.IndexFlatIP(dim)
        index.add(emb)

        tfidf_pack = None
        if self.cfg.hybrid.enable:
            vectorizer = TfidfVectorizer(max_features=self.cfg.hybrid.max_features)
            X = vectorizer.fit_transform(chunks_texts)
            X = sk_normalize(X, norm="l2", axis=1)
            tfidf_pack = {"vectorizer": vectorizer, "X": X}

        doc_dir = self.index_dir / json_path.stem
        doc_dir.mkdir(parents=True, exist_ok=True)
        faiss.write_index(index, str(doc_dir / "faiss.index"))
        with (doc_dir / "meta.pkl").open("wb") as f:
            pickle.dump(chunks_meta, f)
        if tfidf_pack is not None:
            with (doc_dir / "tfidf.pkl").open("wb") as f:
                pickle.dump(tfidf_pack, f)
        with (doc_dir / "info.json").open("w", encoding="utf-8") as f:
            json.dump({"chunks": len(chunks_texts)}, f)

        return doc_dir

    def build_all(self) -> List[Path]:
        parsed_dir = Path(self.cfg.parsed_dir)
        jsons = list(parsed_dir.glob("*.json"))
        if not jsons:
            log.warning("В каталоге %s нет JSON репортов", parsed_dir)
            return []
        out_dirs = []
        for j in tqdm(jsons, desc="Build indices"):
            d = self._build_doc_index(j)
            if d is not None:
                out_dirs.append(d)
        return out_dirs


# ==========================
# Retriever
# ==========================

log = logging.getLogger(__name__)
_client = OpenAI()

UNSUPP_TEMP_MODELS = {"gpt-5", "gpt-5-mini", "gpt-5.1", "gpt-5.1-mini"}

def _chat_create_sanitized(
    model: str,
    messages: List[Dict[str, str]],
    *,
    timeout: Optional[float] = None,
    **kwargs: Any,
):
    # Drop unsupported params for certain models
    if model in UNSUPP_TEMP_MODELS:
        kwargs.pop("temperature", None)
        # Optionally drop other tuning params if your model rejects them:
        # kwargs.pop("top_p", None)
        # kwargs.pop("presence_penalty", None)
        # kwargs.pop("frequency_penalty", None)
    # Remove None values to avoid sending them
    clean_kwargs = {k: v for k, v in kwargs.items() if v is not None}
    return _client.chat.completions.create(
        model=model,
        messages=messages,
        **({"timeout": timeout} if timeout else {}),
        **clean_kwargs,
    )


class RAGRetriever:
    def __init__(self, cfg: RetrieverConfig, rerank: Optional[RerankConfig] = None):
        self.cfg = cfg
        self.client = OpenAI()
        self.rerank = rerank or RerankConfig(enable=False)
        self._cache = {}

    def _ensure_doc_loaded(self, doc_dir: Path):
        key = str(doc_dir.resolve())
        if key in self._cache:
            return
        index = faiss.read_index(str(doc_dir / "faiss.index"))
        with (doc_dir / "meta.pkl").open("rb") as f:
            meta = pickle.load(f)
        tfidf_pack = None
        tfidf_path = doc_dir / "tfidf.pkl"
        if tfidf_path.exists() and self.cfg.hybrid.enable:
            with tfidf_path.open("rb") as f:
                tfidf_pack = pickle.load(f)
        self._cache[key] = (index, meta, tfidf_pack)

    def _embed_query(self, text: str) -> np.ndarray:
        api_key = os.getenv(self.cfg.embed.openai_api_key_env)
        if not api_key:
            raise RuntimeError(f"{self.cfg.embed.openai_api_key_env} is not set")
        resp = self.client.embeddings.create(model=self.cfg.embed.model, input=[text])
        v = np.array([resp.data[0].embedding], dtype="float32")
        v /= (np.linalg.norm(v, axis=1, keepdims=True) + 1e-12)
        return v

    def _score_hybrid(self, doc_dir: Path, q: str, top_k_doc: int = 50) -> List[Dict[str, Any]]:
        vq = self._embed_query(q)
        key = str(doc_dir.resolve())
        index, meta, tfidf_pack = self._cache[key]
        faiss_scores, faiss_ids = index.search(vq, min(top_k_doc, index.ntotal))
        faiss_scores = faiss_scores[0].tolist()
        faiss_ids = faiss_ids[0].tolist()

        lex_scores = None
        if self.cfg.hybrid.enable and tfidf_pack is not None:
            vectorizer = tfidf_pack["vectorizer"]
            X = tfidf_pack["X"]
            q_vec = vectorizer.transform([q])
            q_vec = sk_normalize(q_vec, norm="l2", axis=1)
            lex = (X @ q_vec.T).toarray().ravel()
            lex_scores = [float(lex[i]) for i in faiss_ids]

        hits = []
        for rank, cid in enumerate(faiss_ids):
            item = {
                "_index_dir": doc_dir.name,
                "chunk_id": cid,
                "faiss_score": float(faiss_scores[rank]),
                "text": meta[cid]["text"],
                "page_id": meta[cid]["page_id"],
            }
            if lex_scores is not None:
                item["lex_score"] = float(lex_scores[rank])
                wF = float(self.cfg.hybrid.weight_faiss)
                wL = float(self.cfg.hybrid.weight_lexical)
                item["score"] = wF * item["faiss_score"] + wL * item["lex_score"]
            else:
                item["score"] = item["faiss_score"]
            hits.append(item)

        hits.sort(key=lambda x: x["score"], reverse=True)
        return hits


    def _llm_rerank_scores(self, q: str, texts: List[str]) -> List[float]:
        # sys = ("You are a strict retrieval judge. "
        #        "Rate how useful the text is for answering the query from 0.0 (irrelevant) to 1.0 (perfect). "
        #        "Return ONLY the number.")

        sys = (
            "You are a strict retrieval judge.\n"
            "Rate how useful the text is for answering the query in [0,1].\n"
            "Scale:\n"
            "- 0.8–1.0: explicit answer/numbers/entities from the query.\n"
            "- 0.6–0.8: highly relevant, near-explicit.\n"
            "- 0.4–0.6: thematically related, no direct answer.\n"
            "- 0.2–0.4: weak topical relation.\n"
            "- 0.0–0.2: off-topic. Use 0.0 only if completely unrelated.\n"
            "Return ONLY JSON: {\"score\": number}."
        )
        out: List[float] = []
        print("curr temperature ", self.rerank.temperature)
        import re as _re
        _num = _re.compile(r"[-+]?\d*\.?\d+")
        for t in texts:
            messages = [
                {"role": "system", "content": sys},
                {"role": "user", "content": f"Query:\n{q}\n\nText:\n{t}\n\nScore (0.0..1.0) only:"}
            ]
            try:
                # Uses sanitized helper that strips unsupported params (e.g., 'temperature') for certain models
                r = _chat_create_sanitized(
                    model=self.rerank.llm_model,
                    messages=messages,
                    temperature=self.rerank.temperature,
                    seed = self.rerank.seed + 0,# NEW
                    n= 1
                )
                #print(r)
                content = (r.choices[0].message.content or "").strip()

                m = _num.search(content)
                s = float(m.group(0)) if m else 0.0
            except Exception as e:
                log.warning("LLM rerank failed: %s", e)
                s = 0.0
            out.append(max(0.0, min(1.0, s)))
        return out

    def search_all_docs(self, query: str, top_k: Optional[int] = None) -> List[Dict[str, Any]]:
        top_k = top_k or self.cfg.top_k
        doc_dirs = [p for p in Path(self.cfg.index_dir).iterdir() if p.is_dir()]
        all_hits: List[Dict[str, Any]] = []
        for d in doc_dirs:
            self._ensure_doc_loaded(d)
            doc_hits = self._score_hybrid(d, query, top_k_doc=max(self.rerank.top_k_pre_rerank, top_k))
            for h in doc_hits:
                h["_doc_dir"] = str(d)
            all_hits.extend(doc_hits)

        all_hits.sort(key=lambda x: x["score"], reverse=True)

        if self.rerank.enable and all_hits:
            cand = all_hits[:self.rerank.top_k_pre_rerank]
            texts = [h["text"] for h in cand]
            llm_scores = self._llm_rerank_scores(query, texts)
            for h, s in zip(cand, llm_scores):
                h["llm_score"] = float(s)
                wF = float(self.rerank.weight_faiss)
                wL = float(self.rerank.weight_llm)
                h["final_score"] = wF * h["score"] + wL * s
            cand.sort(key=lambda x: x.get("final_score", x["score"]), reverse=True)
            hits = cand[:top_k]
        else:
            hits = all_hits[:top_k]

        if self.cfg.unique_by_page:
            seen = set()
            uniq = []
            for h in hits:
                key = (h["_index_dir"], h["page_id"])
                if key in seen:
                    continue
                seen.add(key)
                uniq.append(h)
            hits = uniq[:top_k]

        return hits


# ==========================
# QA
# ==========================

def _extract_companies_from_subset(question_text: str, subset_csv_path: Path) -> List[str]:
    try:
        df = pd.read_csv(subset_csv_path)
    except Exception:
        return []
    if "company_name" not in df.columns:
        return []
    names = sorted(df["company_name"].astype(str).unique(), key=len, reverse=True)
    found = []
    s = question_text
    import re as _re
    for nm in names:
        esc = _re.escape(nm)
        pat = rf"{esc}(?:\W|$)"
        if _re.search(pat, s, _re.IGNORECASE):
            found.append(nm)
            s = _re.sub(pat, "", s, flags=_re.IGNORECASE)
    return found

def _format_hits_context(hits: List[Dict[str, Any]], max_ctx_chunks: int = 8, max_snippet_chars: int = 1400) -> str:
    parts = []
    for i, h in enumerate(hits[:max_ctx_chunks], start=1):
        src = h.get("_index_dir") or h.get("doc_name") or "unknown"
        page = h.get("page_id", "?")
        text = (h.get("text") or "").replace("\r", " ").strip()
        if len(text) > max_snippet_chars:
            text = text[:max_snippet_chars] + "…"
        parts.append(f"[{i}] {src} (page {page})\n{text}")
    return "\n\n---\n\n".join(parts) if parts else "(no context)"

def _coerce_answer(ans: str, answer_type: str) -> str:
    at = (answer_type or "str").lower()
    s = (ans or "").strip().strip('"').strip("'")
    if at == "int":
        import re as _re
        m = _re.search(r"[-+]?\d+", s)
        return m.group(0) if m else "N/A"
    if at == "float":
        import re as _re
        m = _re.search(r"[-+]?\d+(?:\.\d+)?", s.replace(",", "."))
        return m.group(0) if m else "N/A"
    return s

from typing import Any, Dict, List
import logging

log = logging.getLogger(__name__)

def llm_answer_from_hits(
    question: str,
    hits: List[Dict[str, Any]],
    answer_type: str = "str",
    model: str = "gpt-4o-mini",
    max_ctx_chunks: int = 8,
    timeout: float = 60.0,
) -> str:
    """
    Call the LLM for a single final value using only the provided context and question.
    """
    client = OpenAI()
    context_block = _format_hits_context(hits, max_ctx_chunks=max_ctx_chunks)

    sys_msg = (
        "You are a precise QA assistant for financial and corporate documents.\n"
        "Use ONLY the provided context passages and the question text.\n"
        "Prefer producing a concrete, best‑effort answer. If any direct or indirect evidence exists in the\n"
        "context or the question, infer the single most plausible value. Output 'N/A' ONLY when the answer\n"
        "cannot be inferred at all from the provided context and question, or when the evidence is mutually\n"
        "contradictory.\n\n"
        "Mandatory decisions:\n"
        "- If the question is multiple-choice or a survey item, selecting exactly ONE of the provided\n"
        "  options is mandatory. Never answer 'N/A' for such questions.\n\n"
        "Required (when applicable) reasoning/actions BEFORE producing the final value:\n"
        "- If the question requires a simple calculation (sum, difference, average, percent, ratio),\n"
        "  do it using numbers present in the context or the question.\n"
        "- You may perform basic arithmetic to compute totals/aggregates when required (e.g., summing\n"
        "  items across rows/periods), strictly using values from the provided context or the question.\n\n"
        "Output format: EXACTLY ONE value, no extra text.\n"
        "- answer_type=str  : short string, no quotes.\n"
        "- answer_type=int  : digits only; no spaces; no separators; round towards zero.\n"
        "- answer_type=float: digits with optional single '.'; no spaces; '.' as decimal separator.\n"
    )
    user_msg = f"answer_type={answer_type}\n\nQuestion:\n{question}\n\nContext:\n{context_block}"

    try:
        resp = _chat_create_sanitized(
            model=model,
            messages=[
                {"role": "system", "content": sys_msg},
                {"role": "user", "content": user_msg},
            ],
            timeout=timeout,
        )
        raw = (resp.choices[0].message.content or "").strip()
    except Exception as e:
        log.warning("LLM answer failed: %s", e)
        raw = "N/A"

    return _coerce_answer(raw, answer_type)

import logging

log = logging.getLogger(__name__)


from typing import Optional
from openai import OpenAI
import logging

log = logging.getLogger(__name__)

def expand_multilang_entities(question: str, *, model: str = "gpt-5", client: Optional[OpenAI] = None) -> str:
    """
    Augment ONLY:
      - Personal names (persons)
      - Company/organization names (official names)
    Exclude:
      - Countries/regions/cities/locations (any GPE)
      - Generic/common terms (roles, categories, legal/finance terms)
    Behavior:
      - Exactly ONE translation in parentheses depending on original:
          * English -> append Russian full name
          * Russian -> append English full name
          * Kazakh -> append Russian full name
      - Do NOT expand/translate legal forms (АО/АОАО/ТОО/JSC/LLP/LLC/etc.). Keep them outside as-is.
      - If a span consists ONLY of legal forms (e.g., 'АО', 'ТОО', 'ООО'), do nothing (no parentheses).
      - Inside parentheses include ONLY the base company/person name (no legal forms).
      - Preserve original text; return original on failure.
    """
    if not question or not question.strip():
        return question

    _client = client or OpenAI()

    sys_msg = (
        "You are a query augmenter for document retrieval.\n"
        "Augment ONLY these entity types:\n"
        "  - Personal names (persons)\n"
        "  - Company/organization names (official names)\n"
        "NEVER augment: countries/regions/cities (any GPE) and generic/common terms.\n"
        "Language rule for a SINGLE appended variant:\n"
        "  - If the original entity is English (Latin), append ONLY the Russian full name.\n"
        "  - If the original is Russian (Cyrillic without Kazakh letters), append ONLY the English full name.\n"
        "  - If the original is Kazakh (Cyrillic with ә, ғ, қ, ң, ө, ұ, ү, һ, і), append ONLY the Russian full name.\n"
        "IMPORTANT:\n"
        "  - Do NOT expand or translate legal forms (e.g., АО, АОАО, ТОО, ООО, ПАО, JSC, PJSC, LLC, LLP, Inc., Ltd.).\n"
        "  - Keep legal forms outside parentheses unchanged.\n"
        "  - If a mention consists ONLY of legal forms (no concrete organization/person name), do nothing.\n"
        "  - Inside parentheses include ONLY the base official name of the person/organization (no legal forms).\n"
        "If an entity is a non-legal-form acronym (e.g., a company shortname), you may expand it to its official base name,\n"
        "but still exclude legal forms from the appended variant.\n"
        "After the FIRST occurrence of each allowed entity, append parentheses with EXACTLY ONE full name per the language rule.\n"
        "No language labels; preserve the original wording; return a single line with the rewritten question only."
    )
    user_msg = f"Question:\n{question}"

    try:
        resp = safe_chat_create(
            _client,
            model=model,
            messages=[
                {"role": "system", "content": sys_msg},
                {"role": "user", "content": user_msg},
            ],
        )
        out = (resp.choices[0].message.content or "").strip()

        # Strip accidental wrapping quotes
        m = re.match(r'^[\'"](.+)[\'"]$', out)
        if m:
            out = m.group(1).strip()

        if not out or out == question or "(" not in out:
            return question

        # --------- Heuristics and filters ----------
        QUOTE_CHARS = '«»"“”„‟’‘`'
        LEGAL_FORMS = (
            "АО", "АОАО", "ТОО", "ООО", "ПАО", "ОАО", "ИП",
            "JSC", "PJSC", "LLC", "LLP", "Inc.", "Corp.", "Ltd.", "PLC",
            "GmbH", "AG", "S.A.", "SAS", "BV", "NV", "SA", "S.p.A."
        )
        # Spelled-out legal forms to strip from inside parentheses
        LEGAL_FORMS_LONG = (
            "Акционерное общество", "Публичное акционерное общество",
            "Общество с ограниченной ответственностью",
            "Товарищество с ограниченной ответственностью",
            "Акционерлік қоғам", "Жауапкершілігі шектеулі серіктестік",
            "Joint Stock Company", "Public Joint Stock Company",
            "Limited Liability Company", "Limited Liability Partnership"
        )
        ORG_KEYWORDS = ("банк", "компания", "корпорация", "холдинг", "завод", "фабрика", "концерн", "предприятие")
        GENERIC_EXCLUDE = {
            "физлиц", "физические лица", "физическое лицо",
            "юрлиц", "юридические лица", "юридическое лицо",
            "работники", "сотрудники", "налоги", "налог",
            "договор", "контракт", "контракты", "акции", "облигации",
        }

        kk_chars = "ӘәҒғҚқҢңӨөҰұҮүҺһІі"
        has_lat = lambda s: bool(re.search(r"[A-Za-z]", s))
        has_cyr = lambda s: bool(re.search(r"[А-Яа-яЁё" + kk_chars + r"]", s))
        has_kk = lambda s: bool(re.search("[" + kk_chars + "]", s))

        def detect_lang(s: str) -> str:
            s = s.strip()
            if has_lat(s) and not has_cyr(s):
                return "en"
            if has_kk(s):
                return "kk"
            if has_cyr(s):
                return "ru"
            return "unknown"

        cap_word = r"[A-ZА-ЯЁ][a-zа-яё\-]+"
        person_like_re = re.compile(rf"\b{cap_word}(?:\s+{cap_word}){{1,2}}\b")
        acronym_re = re.compile(r"\b[А-ЯЁA-Z]{2,}\b")

        def tail_segment(s: str) -> str:
            seg = re.split(r"[,\.;:—\-–\n\t]+", s)
            return seg[-1].strip() if seg else s.strip()

        def has_legal_form(a: str) -> bool:
            s = a.strip()
            return any(tok in s for tok in LEGAL_FORMS) or any(sym in s for sym in "«»\"")

        def has_org_keyword(s: str) -> bool:
            s_low = s.lower()
            return any(kw in s_low for kw in ORG_KEYWORDS)

        def is_person_like(s: str) -> bool:
            return bool(person_like_re.search(s))

        def is_generic(s: str) -> bool:
            t = re.sub(rf"[{re.escape(QUOTE_CHARS)}()]", "", s).strip().lower()
            t = re.sub(r"\s+", " ", t)
            return t in GENERIC_EXCLUDE

        # NEW: detect spans that are ONLY legal forms (e.g., "АО", "ТОО", "ООО", possibly combined with 'и'/'and' etc.)
        LEGAL_SET = {lf.lower().strip(".") for lf in LEGAL_FORMS}
        LEGAL_LONG_SET = {lf.lower() for lf in LEGAL_FORMS_LONG}
        CONNECTORS = {"и", "или", "и/или", "and", "or", "&", "/", "-", "—"}

        def is_only_legal_forms_span(s: str) -> bool:
            tokens = re.findall(r"[A-Za-zА-Яа-яЁё\.\-]+", s)
            if not tokens:
                return False
            for t in tokens:
                tl = t.strip().lower().strip(".")
                if not tl or tl in CONNECTORS:
                    continue
                if tl in LEGAL_SET or tl in LEGAL_LONG_SET:
                    continue
                return False
            return True

        # Strip legal forms from inside parentheses (begin/end, multiple passes)
        LF_PAT = r"(?:%s)" % "|".join(
            [re.escape(x) for x in LEGAL_FORMS]
            + [x.replace(" ", r"\s+") for x in LEGAL_FORMS_LONG]
        )
        def strip_legal_forms_inside(s: str) -> str:
            s2 = s.strip()
            if not s2:
                return s2
            punct = r"[ ,\.;:—\-–\"«»“”„‟’‘`]+"
            while True:
                prev = s2
                s2 = re.sub(rf"^(?:{LF_PAT})(?:\.?{punct})+", "", s2, flags=re.IGNORECASE)
                s2 = re.sub(rf"(?:{punct}\.?)?(?:{LF_PAT})\s*$", "", s2, flags=re.IGNORECASE)
                s2 = s2.strip()
                if s2 == prev:
                    break
            return s2

        # 1) Drop likely location augmentations after prepositions with single-word capitalized names.
        PREPS = r"(?:в|во|на|из|с|со|к|ко|от|по|о|об|обо|у|до|при|над|под|перед|за|между|среди)"
        pre_loc_rx = re.compile(
            rf"(?P<prep>\b{PREPS}\b)\s+(?P<name>[A-ZА-ЯЁ][A-Za-zА-Яа-яЁё\-]+)\s*\((?P<inside>[^)]*)\)"
        )
        def drop_location_like(m: re.Match) -> str:
            name = m.group("name")
            if re.fullmatch(r"[A-ZА-ЯЁ]{2,}", name) or has_legal_form(name) or has_org_keyword(name) or is_person_like(name):
                return m.group(0)
            return f"{m.group('prep')} {name}"

        cleaned = pre_loc_rx.sub(drop_location_like, out)

        # 2) General post-filter: keep only persons and companies; enforce single-language rule and strip legal forms inside
        any_anchor_rx = re.compile(r"(?P<a>[^\(\)\n]{1,160}?)\s*\((?P<i>[^)]{1,240})\)")

        def keep_only_valid(m: re.Match) -> str:
            anchor_full = m.group("a")
            anchor = tail_segment(anchor_full)
            inside = (m.group("i") or "").strip()

            # If anchor is only legal forms, keep as-is (no augmentation)
            if is_only_legal_forms_span(anchor):
                return anchor_full.rstrip()

            # Drop generics
            if is_generic(anchor):
                return anchor_full.rstrip()

            is_orgish = has_legal_form(anchor) or has_org_keyword(anchor) or acronym_re.fullmatch(anchor.strip())
            is_person = is_person_like(anchor)

            if not (is_orgish or is_person):
                return anchor_full.rstrip()

            # Exactly one variant (no pipes)
            if "|" in inside:
                return anchor_full.rstrip()

            # Language mapping
            src_lang = detect_lang(anchor)
            tgt = "ru" if src_lang in ("en", "kk") else ("en" if src_lang == "ru" else "unknown")
            if tgt == "ru" and not has_cyr(inside):
                return anchor_full.rstrip()
            if tgt == "en" and not has_lat(inside):
                return anchor_full.rstrip()
            if tgt == "unknown":
                return anchor_full.rstrip()

            # Strip legal forms inside parentheses (companies only)
            new_inside = strip_legal_forms_inside(inside) if is_orgish else inside
            if not new_inside:
                return anchor_full.rstrip()

            # Rebuild with sanitized inside; anchor (with legal forms) stays unchanged
            return f"{anchor_full.strip()} ({new_inside})"

        cleaned = any_anchor_rx.sub(keep_only_valid, cleaned)

        if "(" not in cleaned or cleaned == question:
            return question
        return cleaned
    except Exception as e:
        log.warning("expand_multilang_entities failed: %s", e)
        return question

from typing import Any, Dict, Optional


def _extract_json_obj(text: str) -> Optional[Dict[str, Any]]:
    """
    Best‑effort extraction of a single JSON object from an LLM reply.
    Returns a dict or None.
    """
    if not text:
        return None

    def _strip_fences(s: str) -> str:
        s = s.strip()
        # ```json ... ``` or ``` ... ```
        if s.startswith("```"):
            s = re.sub(r"^```[a-zA-Z0-9]*\s*", "", s, count=1, flags=re.DOTALL)
            s = re.sub(r"\s*```$", "", s, count=1, flags=re.DOTALL)
        # Inline code blocks
        s = s.strip("` \n\r\t")
        return s.strip()

    def _try_load(s: str) -> Optional[Any]:
        try:
            return json.loads(s)
        except Exception:
            return None

    def _repair_json(s: str) -> str:
        # Replace Python literals with JSON
        s = re.sub(r"\bTrue\b", "true", s)
        s = re.sub(r"\bFalse\b", "false", s)
        s = re.sub(r"\bNone\b", "null", s)
        # Remove trailing commas before } or ]
        s = re.sub(r",\s*([}\]])", r"\1", s)
        # Naive single‑quote to double‑quote conversion
        # (kept simple for typical LLM outputs)
        if '"' not in s and "'" in s:
            s = s.replace("'", '"')
        return s

    def _first_json_object_span(s: str) -> Optional[str]:
        # Find first balanced {...} while respecting string quotes and escapes
        i = s.find("{")
        while i != -1:
            depth = 0
            in_str = False
            esc = False
            quote = ""
            for j in range(i, len(s)):
                ch = s[j]
                if in_str:
                    if esc:
                        esc = False
                    elif ch == "\\":
                        esc = True
                    elif ch == quote:
                        in_str = False
                else:
                    if ch in ("'", '"'):
                        in_str = True
                        quote = ch
                    elif ch == "{":
                        depth += 1
                    elif ch == "}":
                        depth -= 1
                        if depth == 0:
                            return s[i:j+1]
            # try next '{'
            i = s.find("{", i + 1)
        return None

    raw = _strip_fences(text)

    # 1) Direct parse
    obj = _try_load(raw)
    if isinstance(obj, dict):
        return obj
    if isinstance(obj, list) and obj and isinstance(obj[0], dict):
        return obj[0]

    # 2) Extract first JSON object substring and parse
    span = _first_json_object_span(raw)
    if span:
        obj = _try_load(span)
        if isinstance(obj, dict):
            return obj

        repaired = _repair_json(span)
        obj = _try_load(repaired)
        if isinstance(obj, dict):
            return obj

    # 3) As a last resort, repair whole string and parse
    repaired = _repair_json(raw)
    obj = _try_load(repaired)
    if isinstance(obj, dict):
        return obj
    if isinstance(obj, list) and obj and isinstance(obj[0], dict):
        return obj[0]

    return None


from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openai import OpenAI

def _load_src_filename_from_meta_resolve(parsed_dir: Path, doc_ref: Optional[str]) -> str:
    """
    Best-effort resolve a human-readable source filename for a hit.
    Accepts either an index dir path, a doc name, or None.
    Returns a short name like 'mydoc.pdf' or 'unknown' on failure.
    """
    try:
        if not doc_ref:
            return "unknown"
        p = Path(str(doc_ref))
        # If this looks like an index dir with info.json, try to read original filename
        if p.is_dir() and (p / "info.json").exists():
            try:
                info = json.loads((p / "info.json").read_text(encoding="utf-8"))
                for k in ("source_pdf", "source_file", "document_name", "pdf_name"):
                    v = info.get(k)
                    if isinstance(v, str) and v.strip():
                        return Path(v).name
            except Exception:
                pass
            return p.name
        # Try to infer from parsed dir presence
        cand = parsed_dir / f"{p.name}.json"
        if cand.exists():
            return f"{p.name}.pdf"
        return p.name
    except Exception:
        return "unknown"


def _llm_answer_with_support(
    client: OpenAI,
    model: str,
    question: str,
    answer_type: str,
    hits: List[Dict[str, Any]],
    max_ctx_chunks: int = 8,
    max_snippet_chars: int = 1800,
) -> Tuple[Optional[str], Optional[int], Optional[str]]:
    """
    Ask the LLM to answer using ONLY provided passages and return:
    (answer_string, support_idx_1based, raw_llm_output)
    """
    # Prepare numbered context [1], [2], ...
    ctx_parts: List[str] = []
    for i, h in enumerate(hits[:max_ctx_chunks], start=1):
        src = h.get("source") or h.get("_index_dir") or h.get("doc_name") or "unknown"
        page = h.get("page_id", "?")
        text = (h.get("text") or "").strip().replace("\r", " ")
        if len(text) > max_snippet_chars:
            text = text[:max_snippet_chars] + "…"
        ctx_parts.append(f"[{i}] {src} (page {page})\n{text}")
    context_block = "\n\n---\n\n".join(ctx_parts) if ctx_parts else "(no context)"

    at = (answer_type or "str").strip().lower()
    if at not in ("str", "int", "float"):
        at = "str"

    sys_msg = (
        "You are a precise QA assistant for financial and corporate documents.\n"
        "Use ONLY the provided passages. Produce a JSON object with two fields:\n"
        '  {\"answer\": <value>, \"support\": <passage_index>}\n'
        "- support is the index [1..N] of the SINGLE passage that best supports the answer.\n"
        "- For int: digits only. For float: '.' decimal separator. No units.\n"
        "- For str: short string without quotes.\n"
        "Return ONLY the JSON object."
    )
    user_msg = (
        f"question: {question}\n"
        f"answer_type: {at}\n\n"
        f"PASSAGES:\n{context_block}\n\n"
        "Output JSON now:"
    )

    params: Dict[str, Any] = {
        "model": model,
        "messages": [
            {"role": "system", "content": sys_msg},
            {"role": "user", "content": user_msg},
        ],
        "temperature": 0,
    }

    try:
        resp = client.chat.completions.create(**params)
    except Exception:
        params.pop("temperature", None)
        resp = client.chat.completions.create(**params)

    raw = (resp.choices[0].message.content or "").strip()
    obj = _extract_json_obj(raw) or {}

    ans = obj.get("answer", None)
    ans_str = None if ans is None else str(ans).strip()

    sup = obj.get("support", None)
    try:
        sup_idx = int(sup) if sup is not None else None
    except Exception:
        sup_idx = None

    # Clamp to available passages
    max_allowed = min(len(hits), max_ctx_chunks)
    if sup_idx is not None and (sup_idx < 1 or sup_idx > max_allowed):
        sup_idx = None

    return ans_str, sup_idx, raw

# python
from typing import Any, Dict, List, Optional


def _guess_support_by_value(hits: List[Dict[str, Any]], value: Optional[str]) -> Optional[int]:
    """
    Best‑effort guess of the supporting passage index (1‑based) based on the final coerced value.
    - For numeric answers: prefer exact string occurrence; otherwise choose the chunk whose
      closest number is nearest to the target value (absolute and relative distance).
    - For string answers: prefer exact phrase; otherwise rank by token overlap.
    Returns None if no reasonable guess can be made.
    """
    if not hits or not value:
        return None

    val = value.strip()
    if not val or val.upper() == "N/A":
        return None

    # --- helpers ---
    def _norm_text(s: str) -> str:
        s = (s or "").replace("\r", " ")
        s = re.sub(r"\s+", " ", s)
        return s.strip()

    def _parse_float(s: str) -> Optional[float]:
        try:
            # Normalize thousands/decimal separators: handle "1 234,56" or "1,234.56"
            s2 = s.replace("\u00A0", " ")
            s2 = s2.replace(" ", "")
            # If both '.' and ',' present, assume ',' is thousands sep -> drop ',' keep '.'
            if "." in s2 and "," in s2:
                s2 = s2.replace(",", "")
            else:
                # If only ',' present, treat it as decimal separator
                s2 = s2.replace(",", ".")
            return float(s2)
        except Exception:
            return None

    def _is_numeric_answer(ans: str) -> bool:
        return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", ans))

    # Regex to extract numbers with optional thousands and decimals
    num_rx = re.compile(r"[-+]?\d{1,3}(?:[ \u00A0.,]\d{3})*(?:[.,]\d+)?|[-+]?\d+(?:[.,]\d+)?")

    # --- try numeric first ---
    if _is_numeric_answer(val):
        target = _parse_float(val)
        if target is None:
            # Fallback to string mode
            pass
        else:
            best_idx = None
            best_score = float("inf")
            exact_match_idx = None

            for i, h in enumerate(hits):
                txt = _norm_text(h.get("text", ""))
                if not txt:
                    continue

                # Prefer exact literal occurrence of the answer
                if val in txt:
                    # First exact match wins decisively
                    exact_match_idx = i
                    break

                # Otherwise compute nearest numeric distance in this chunk
                distances: List[float] = []
                for m in num_rx.finditer(txt):
                    f = _parse_float(m.group(0))
                    if f is None:
                        continue
                    # Combined absolute and relative distance
                    abs_d = abs(f - target)
                    rel_d = abs_d / (abs(target) + 1e-9)
                    distances.append(abs_d + 0.5 * rel_d)

                if not distances:
                    continue

                score = min(distances)
                if score < best_score:
                    best_score = score
                    best_idx = i

            if exact_match_idx is not None:
                return exact_match_idx + 1

            # Apply a loose sanity threshold: accept only if the score is reasonably small
            if best_idx is not None and best_score < 0.25 or (abs(target) >= 10 and best_score < 1.5):
                return best_idx + 1

            # No confident numeric support
            return None

    # --- string matching mode ---
    phrase = val.lower()
    phrase = phrase.strip('"').strip("'")
    if not phrase:
        return None

    tokens = [t for t in re.split(r"\W+", phrase) if len(t) >= 2]
    if not tokens:
        # Fallback to exact phrase search if tokens are too short
        tokens = [phrase]

    best_idx = None
    best_score = -1.0

    for i, h in enumerate(hits):
        txt_raw = h.get("text", "") or ""
        txt = _norm_text(txt_raw).lower()
        if not txt:
            continue

        # Exact phrase match gets a big boost
        score = 0.0
        if phrase and phrase in txt:
            score += 2.5

        # Token overlap
        present = sum(1 for t in tokens if t in txt)
        score += present / max(1.0, len(tokens))

        if score > best_score:
            best_score = score
            best_idx = i

    # Require a minimal score to avoid random selection
    if best_idx is not None and best_score >= 0.5:
        return best_idx + 1

    return None

# python

import re
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

try:
    from openai import OpenAI
except Exception:
    OpenAI = None  # type: ignore

# Logger (safe default if not globally set)
log = logging.getLogger("rag_pipeline_integrated")
if not log.handlers:
    h = logging.StreamHandler()
    h.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(name)s | %(message)s"))
    log.addHandler(h)
    log.setLevel(logging.INFO)

# -------------------------
# Robust PDF path resolving
# -------------------------
_POSSIBLE_META_KEYS = (
    "pdf_path", "source_pdf", "src_pdf_path", "original_pdf", "src_path", "source", "document_path", "source_path"
)
_POSSIBLE_NAME_KEYS = ("doc_name", "document_name", "source", "_index_dir")


def _iter_candidate_stems_from_hit(hit: Dict[str, Any]) -> List[str]:
    stems: List[str] = []
    for k in _POSSIBLE_NAME_KEYS + ("source_path",):
        v = hit.get(k)
        if not v:
            continue
        p = Path(str(v))
        if p.stem:
            stems.append(p.stem)
        if p.name and p.name not in stems:
            stems.append(p.name)
    # Dedup preserve order
    uniq, seen = [], set()
    for s in stems:
        if s and s not in seen:
            uniq.append(s); seen.add(s)
    return uniq


def _resolve_from_info_json(idx_dir: Path, roots: Sequence[Path]) -> Optional[Path]:
    info_path = idx_dir / "info.json"
    if not info_path.exists():
        return None
    try:
        info = json.loads(info_path.read_text(encoding="utf-8"))
    except Exception as e:
        log.debug("Failed to read %s: %s", info_path, e)
        return None

    # 1) direct meta keys
    for k in _POSSIBLE_META_KEYS:
        v = info.get(k)
        if not v:
            continue
        p = Path(str(v))
        if p.is_absolute() and p.exists():
            return p
        for r in roots:
            cand = (r / p).resolve()
            if cand.exists():
                return cand
        if p.suffix.lower() != ".pdf":
            for r in roots:
                cand = (r / f"{p.stem}.pdf").resolve()
                if cand.exists():
                    return cand

    # 2) doc_stem
    stem = (info.get("doc_stem") or "").strip()
    if stem:
        for r in roots:
            cand = (r / f"{stem}.pdf").resolve()
            if cand.exists():
                return cand
        # Deep search if needed
        for r in roots:
            m = list(r.rglob(f"{stem}.pdf"))
            if m:
                return m[0]
    return None


def _glob_first_pdf(roots: Sequence[Path], stem: str) -> Optional[Path]:
    # Try exact file: **/stem.pdf
    for r in roots:
        for cand in r.rglob(f"{stem}.pdf"):
            return cand
    # Try prefix: **/stem*.pdf
    for r in roots:
        for cand in r.rglob(f"{stem}*.pdf"):
            return cand
    return None


def _resolve_pdf_path_from_hit(hit: Dict[str, Any], pdf_roots: Sequence[Path]) -> Optional[Path]:
    """
    Resolve absolute PDF path for a retrieval hit using:
    1) explicit fields in the hit,
    2) index_dir/info.json (if available),
    3) name-based glob search under provided roots.
    """
    roots = [Path(r).resolve() for r in pdf_roots if r]
    # 1) direct fields from the hit
    for k in _POSSIBLE_META_KEYS:
        v = hit.get(k)
        if not v:
            continue
        p = Path(str(v))
        if p.is_absolute() and p.exists():
            return p
        for r in roots:
            cand = (r / p).resolve()
            if cand.exists():
                return cand
        if p.suffix.lower() != ".pdf":
            for r in roots:
                cand = (r / f"{p.stem}.pdf").resolve()
                if cand.exists():
                    return cand

    # 2) info.json inside index dir (if provided)
    idx_dir_val = hit.get("_index_dir")
    if idx_dir_val:
        idx_dir = Path(str(idx_dir_val))
        if idx_dir.exists():
            resolved = _resolve_from_info_json(idx_dir, roots)
            if resolved:
                return resolved

    # 3) name-based deep search
    for stem in _iter_candidate_stems_from_hit(hit):
        cand = _glob_first_pdf(roots, stem)
        if cand:
            return cand

    return None


# -------------------------
# PDF page rendering
# -------------------------
def _render_page_to_png_b64(pdf_path: Path, page_num_1b: int, dpi: int = 200) -> Optional[str]:
    """
    Render 1-based PDF page to base64 PNG and return the string.
    """
    doc = None
    try:
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        log.warning("Failed to open PDF: %s (%s)", pdf_path, e)
        return None
    try:
        total = doc.page_count
        if page_num_1b < 1 or page_num_1b > total:
            log.warning("Page out of range: %s (page=%s, total=%s)", pdf_path, page_num_1b, total)
            return None
        mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
        page = doc.load_page(page_num_1b - 1)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        png_bytes = pix.tobytes("png")
        b64 = base64.b64encode(png_bytes).decode("ascii")
        log.info("Rendered page image: file=%s page=%s dpi=%s size=%dB", pdf_path.name, page_num_1b, dpi, len(png_bytes))
        return b64
    except Exception as e:
        log.warning("Failed to render page %s of %s: %s", page_num_1b, pdf_path, e)
        return None
    finally:
        try:
            if doc:
                doc.close()
        except Exception:
            pass


# -------------------------
# LLM call (multimodal)
# -------------------------
# python
import logging
from typing import Any, Dict, List, Optional, Sequence, Tuple

log = logging.getLogger("rag_pipeline_integrated")

def _safe_text_from_hit(h: Dict[str, Any], max_chars: int = 1200) -> str:
    # Try common fields for retrieved chunk text
    for k in ("chunk_text", "text", "snippet", "content", "body"):
        v = h.get(k)
        if isinstance(v, str) and v.strip():
            s = v.strip()
            return s if len(s) <= max_chars else s[: max_chars - 3] + "..."
    return ""

def _build_messages(question: str, answer_type: str, hits: List[Dict[str, Any]], image_b64: Optional[str]) -> List[Dict[str, Any]]:
    system_prompt = (
        "You are a factual QA assistant. Answer concisely using ONLY the provided snippets and optional page image. "
        "Return the best short answer matching the requested answer_type. "
        "If unknown from the provided context, return the string 'unknown'. "
        "Also return a 1-based support index pointing to the most relevant snippet."
    )

    # Prepare textual context
    parts: List[Dict[str, Any]] = []
    ctx_lines: List[str] = []
    for i, h in enumerate(hits, start=1):
        doc = h.get("doc_name") or h.get("document_name") or h.get("source") or h.get("doc_stem") or h.get("_index_dir") or ""
        page = h.get("page_num") or h.get("page") or h.get("page_number") or ""
        txt = _safe_text_from_hit(h)
        ctx_lines.append(f"[{i}] doc={doc} page={page}\n{txt}".strip())
    ctx_block = "\n\n".join(ctx_lines) if ctx_lines else "(no snippets)"

    # Main user instruction (text part)
    text_user = (
        f"Question: {question}\n"
        f"Answer type: {answer_type}\n\n"
        f"Snippets:\n{ctx_block}\n\n"
        f"Instructions: Use only the snippets and the optional image. "
        f"Return JSON with fields 'answer' and 'support'. 'support' is the 1-based index of the best snippet."
    )
    parts.append({"type": "text", "text": text_user})

    # Optional image part
    if image_b64:
        parts.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/png;base64,{image_b64}"}
        })

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": parts},
    ]
    return messages

def _call_chat_json(client: Any, model: str, messages: List[Dict[str, Any]], max_tokens: int = 128) -> str:
    resp = client.chat.completions.create(
        model=model,
        messages=messages,
        temperature=0,
        max_tokens=max_tokens,
        response_format={"type": "json_object"},
    )
    return (resp.choices[0].message.content or "").strip()

def _looks_like_refusal(s: str) -> bool:
    t = s.lower()
    return ("i'm sorry" in t) or ("i am sorry" in t) or ("can't assist" in t) or ("cannot assist" in t)

def _parse_json_answer(s: str) -> Tuple[Optional[str], Optional[int]]:
    try:
        obj = json.loads(s)
        ans = obj.get("answer")
        sup = obj.get("support")
        if isinstance(sup, str) and sup.isdigit():
            sup = int(sup)
        if isinstance(sup, (int, float)):
            sup = int(sup)
        if ans is not None and not isinstance(ans, str):
            ans = str(ans)
        return ans, sup if isinstance(sup, int) else None
    except Exception:
        return None, None
# python
# --- Drop-in: robust JSON-Only LLM call with ctx limiting and image support ---

import logging
from typing import Any, Dict, List, Optional, Tuple

log = logging.getLogger("rag_pipeline_integrated")

def _safe_text_from_hit(h: Dict[str, Any], max_chars: int) -> str:
    for k in ("chunk_text", "text", "snippet", "content", "body"):
        v = h.get(k)
        if isinstance(v, str) and v.strip():
            s = v.strip().replace("\r", " ")
            return s if len(s) <= max_chars else s[: max_chars - 1] + "…"
    return ""

def _build_messages(
    question: str,
    answer_type: str,
    hits: List[Dict[str, Any]],
    image_b64: Optional[str],
    max_ctx_chunks: int,
    max_snippet_chars: int,
) -> List[Dict[str, Any]]:
    system_prompt = (
        "You are a factual QA assistant. Use ONLY the provided snippets and optional page image.\n"
        "Return a JSON object with fields 'answer' and 'support'.\n"
        "- 'answer' must match the requested answer_type (str|int|float). If unknown from the snippets/image, return 'unknown'.\n"
        "- 'support' is the 1-based index of the SINGLE most relevant snippet."
    )

    # Limit context to max_ctx_chunks
    limited_hits = hits[: max(0, int(max_ctx_chunks or 0))]
    if not limited_hits and hits:
        limited_hits = hits[:1]

    # Build user content (multimodal)
    parts: List[Dict[str, Any]] = []
    lines: List[str] = []
    for i, h in enumerate(limited_hits, start=1):
        doc = h.get("doc_name") or h.get("document_name") or h.get("source") or h.get("_index_dir") or "doc"
        page = h.get("page_id") or h.get("page_number") or h.get("page") or "?"
        txt = _safe_text_from_hit(h, max_snippet_chars)
        lines.append(f"[{i}] {doc} (page {page})\n{txt}")
    snippets_block = "\n\n".join(lines) if lines else "(no snippets)"

    user_text = (
        f"Question: {question}\n"
        f"Answer type: {answer_type}\n\n"
        f"Snippets:\n{snippets_block}\n\n"
        "Respond with a JSON object only."
    )
    parts.append({"type": "text", "text": user_text})
    if image_b64:
        parts.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}})

    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": parts},
    ]

def _call_chat_json(client: Any, model: str, messages: List[Dict[str, Any]], max_tokens: int = 128) -> str:
    resp = client.chat.completions.create(
        model=model,
        messages=messages,
        temperature=0,
        max_tokens=max_tokens,
        response_format={"type": "json_object"},
    )
    return (resp.choices[0].message.content or "").strip()

def _looks_like_refusal(s: str) -> bool:
    t = (s or "").lower()
    return ("i'm sorry" in t) or ("i am sorry" in t) or ("can't assist" in t) or ("cannot assist" in t)

def _parse_json_answer(s: str) -> Tuple[Optional[str], Optional[int]]:
    try:
        obj = json.loads(s)
    except Exception:
        # Try to salvage a JSON region
        s_idx, e_idx = s.find("{"), s.rfind("}")
        if s_idx >= 0 and e_idx > s_idx:
            try:
                obj = json.loads(s[s_idx:e_idx + 1])
            except Exception:
                return None, None
        else:
            return None, None
    ans = obj.get("answer", None)
    sup = obj.get("support", None)
    if isinstance(ans, (int, float)):
        ans = str(ans)
    elif ans is not None and not isinstance(ans, str):
        ans = str(ans)
    if isinstance(sup, str) and sup.isdigit():
        sup = int(sup)
    if isinstance(sup, (int, float)):
        sup = int(sup)
    else:
        sup = None
    return ans, sup

def _llm_answer_with_support(
    client: Any,
    model: str,
    question: str,
    answer_type: str,
    hits: List[Dict[str, Any]],
    max_ctx_chunks: int = 8,
    max_snippet_chars: int = 1800,
    image_b64: Optional[str] = None,
) -> Tuple[Optional[str], Optional[int], Optional[str]]:
    # Primary attempt
    messages = _build_messages(question, answer_type, hits, image_b64, max_ctx_chunks, max_snippet_chars)
    log.info("LLM request: model=%s | image=%s | ctx=%d", model, "yes" if image_b64 else "no", max_ctx_chunks)
    try:
        raw = _call_chat_json(client, model, messages)
        log.info("LLM raw output: %s", raw)
        if _looks_like_refusal(raw):
            raise RuntimeError("refusal_detected")
        ans, sup = _parse_json_answer(raw)
        return ans, sup, raw
    except Exception as e:
        log.warning("Primary LLM call failed/refused (%s). Retrying with fallback.", e)

    # Fallback 1: mini model, same messages
    fallback_model = "gpt-4o-mini"
    try:
        raw = _call_chat_json(client, fallback_model, messages)
        log.info("LLM raw output (fallback-1): %s", raw)
        if _looks_like_refusal(raw):
            raise RuntimeError("refusal_detected")
        ans, sup = _parse_json_answer(raw)
        return ans, sup, raw
    except Exception as e:
        log.warning("Fallback-1 failed/refused (%s). Retrying text-only.", e)

    # Fallback 2: drop image (in case image triggered safety)
    try:
        messages_text_only = _build_messages(question, answer_type, hits, None, max_ctx_chunks, max_snippet_chars)
        raw = _call_chat_json(client, fallback_model, messages_text_only)
        log.info("LLM raw output (fallback-2): %s", raw)
        if _looks_like_refusal(raw):
            raise RuntimeError("refusal_detected")
        ans, sup = _parse_json_answer(raw)
        return ans, sup, raw
    except Exception as e:
        log.error("All LLM attempts failed/refused: %s", e)
        return None, None, None
# -------------------------
# Utility coercion and support guess (local, minimal)
# -------------------------
_num_re = re.compile(r"[-+]?\d+(?:[\.,]\d+)?")

def _coerce_answer(raw: Optional[str], atype: str) -> Optional[Any]:
    if raw is None:
        return None
    at = (atype or "str").strip().lower()
    if at == "int":
        m = _num_re.search(raw)
        if not m:
            return None
        try:
            return int(float(m.group(0).replace(",", ".").replace(" ", "")))
        except Exception:
            return None
    if at == "float":
        m = _num_re.search(raw)
        if not m:
            return None
        try:
            return float(m.group(0).replace(",", ".").replace(" ", ""))
        except Exception:
            return None
    return raw.strip() if isinstance(raw, str) else str(raw)


def _guess_support_by_value(hits: List[Dict[str, Any]], coerced: Any) -> Optional[int]:
    if coerced is None:
        return None
    needle = str(coerced)
    for i, h in enumerate(hits, start=1):
        t = (h.get("text") or "")
        if needle and needle in t:
            return i
    return None


# -------------------------
# Main QA entry point
# -------------------------
def answer_questions_from_xlsx(
    xlsx_path,
    retriever,
    parsed_dir,
    llm_model: str = "gpt-4o",  # vision-capable
    top_k_per_doc: int = 20,
    chunks_in_output: int = 1,
    question_prep=None,
    client: Optional[Any] = None,
    debug: bool = True,
    debug_snippet_chars: int = 220,
    debug_print_top_k: int = 10,
    debug_csv_path=None,
    debug_xlsx_path=None,
    pdf_roots: Optional[Sequence[Path]] = None,  # NEW: where PDFs can be found
) -> List[Dict[str, Any]]:
    import pandas as pd
    from tqdm import tqdm

    if client is None:
        if OpenAI is None:
            raise RuntimeError("OpenAI client is not available; install/import openai.")
        client = OpenAI()

    # Prepare search roots
    roots = [Path(r).resolve() for r in (pdf_roots or []) if r]
    if not roots:
        pd_path = Path(parsed_dir).resolve()
        roots = [pd_path, pd_path.parent, Path.cwd().resolve()]
        log.info("pdf_roots not provided; using fallbacks: %s", [str(r) for r in roots])

    # Load questions
    df = pd.read_excel(xlsx_path)
    cols = {c.lower(): c for c in df.columns}
    for required in ("id", "full_question", "answer_type"):
        if required not in cols:
            raise ValueError(f"Missing required column: {required}")

    results: List[Dict[str, Any]] = []
    csv_rows: List[Dict[str, Any]] = []

    log.info("Starting QA on workbook: %s (rows=%d)", xlsx_path, len(df))

    for _, row in tqdm(df.iterrows(), total=len(df), desc="Answers"):
        qid = int(row[cols["id"]])
        qtext = str(row[cols["full_question"]]).strip()
        atype = str(row[cols["answer_type"]]).strip().lower()
        if question_prep and getattr(question_prep, "prepare", None):
            try:
                qtext = question_prep.prepare(qtext) or qtext
            except Exception:
                pass

        log.info("Q%04d | type=%s | %s", qid, atype, qtext)

        # Retrieve candidates
        hits = retriever.search_all_docs(qtext, top_k=top_k_per_doc) or []

        # Log top-K hits
        if debug and hits:
            for i, h in enumerate(hits[:debug_print_top_k], start=1):
                src = h.get("source") or h.get("_index_dir") or h.get("doc_name") or "unknown"
                page = h.get("page_id") or h.get("page_number") or h.get("page") or "?"
                text = (h.get("text") or "").replace("\r", " ").strip()
                if len(text) > debug_snippet_chars:
                    text = text[:debug_snippet_chars] + " …"
                score_fields = ("score", "faiss_score", "bm25", "lexical", "dense", "rerank_score", "llm_score")
                score_info = ", ".join(
                    f"{k}={h.get(k):.4f}" if isinstance(h.get(k), (int, float)) else f"{k}={h.get(k)}"
                    for k in score_fields
                    if h.get(k) is not None
                )
                log.info("Q%04d | hit#%d | %s (page %s) | %s\n%s", qid, i, src, page, score_info, text)

        # Render image for top hit
        image_b64: Optional[str] = None
        image_info = ""
        image_reason = "ok"
        if hits:
            top_hit = hits[0]
            pdf_path = _resolve_pdf_path_from_hit(top_hit, roots)
            if not pdf_path:
                image_reason = "PDF not resolved"
            else:
                try:
                    page_num = int(top_hit.get("page_id") or top_hit.get("page") or top_hit.get("page_number") or 1)
                except Exception:
                    page_num = 1
                image_b64 = _render_page_to_png_b64(pdf_path, page_num, dpi=200)
                image_info = f"{pdf_path.name}#p{page_num}"
                if not image_b64:
                    image_reason = "render failed"
        else:
            image_reason = "no hits"

        if image_b64:
            log.info("Q%04d | image attached (%s)", qid, image_info)
        else:
            log.info("Q%04d | image skipped (%s)", qid, image_reason)

        # Ask LLM (pass image_b64)
        raw_answer_str, support_1b, raw_llm = _llm_answer_with_support(
            client=client,
            model=llm_model,  # ensure vision-capable, e.g., gpt-4o
            question=qtext,
            answer_type=atype,
            hits=hits,
            max_ctx_chunks=min(8, max(1, top_k_per_doc)),
            image_b64=image_b64,
        )

        # Coerce type
        coerced = _coerce_answer(raw_answer_str, atype)

        # Choose supporting passage
        if support_1b is None:
            support_1b = _guess_support_by_value(hits, coerced) or (1 if hits else None)

        log.info(
            "Q%04d | model=%s | answer_raw=%s | answer=%s | support=%s",
            qid, llm_model, raw_answer_str, coerced, support_1b
        )

        # Build output (exactly one relevant chunk)
        rels: List[Dict[str, Any]] = []
        if hits and support_1b is not None and 1 <= support_1b <= len(hits):
            chosen = hits[support_1b - 1]
            page_no = int(chosen.get("page_id") or chosen.get("page_number") or chosen.get("page") or -1)
            # Prefer hit doc_name; if missing, try to resolve to actual PDF name
            src_name = chosen.get("doc_name")
            if not src_name:
                resolved_pdf = _resolve_pdf_path_from_hit(chosen, roots)
                src_name = resolved_pdf.name if resolved_pdf else (chosen.get("source") or "unknown.pdf")
            rels.append({
                "document_name": src_name,
                "page_number": page_no,
            })
            log.info("Q%04d | chosen chunk -> %s (page %s)", qid, src_name, page_no)
        else:
            log.info("Q%04d | no supporting chunk determined", qid)

        # Debug row
        if debug:
            def _snip(x: Optional[str], n: int = 1000) -> str:
                if not x:
                    return ""
                x = x.replace("\r", " ").strip()
                return x[:n] + (" …" if len(x) > n else "")

            topk_view = []
            for i, h in enumerate(hits[:debug_print_top_k], start=1):
                topk_view.append({
                    "rank": i,
                    "doc": h.get("_index_dir") or h.get("doc_name"),
                    "page": h.get("page_id") or h.get("page_number") or h.get("page"),
                    "offset": h.get("offset"),
                    "score": h.get("score"),
                    "rerank_score": h.get("rerank_score"),
                    "text_snippet": _snip(h.get("text"), 200),
                })

            csv_rows.append({
                "question_id": qid,
                "question": qtext,
                "answer_type": atype,
                "model": llm_model,
                "image": image_info if image_b64 else "",
                "answer_raw": raw_answer_str or "",
                "answer": coerced if coerced is not None else "",
                "support_idx": support_1b or "",
                "support_doc": (rels[0]["document_name"] if rels else ""),
                "support_page": (rels[0]["page_number"] if rels else ""),
                "llm_raw": _snip(raw_llm, 2000),
                "topk_json": json.dumps(topk_view, ensure_ascii=False),
            })

        results.append({
            "question_id": qid,
            "relevant_chunks": rels,  # exactly one
            "answer": (coerced if coerced is not None else ("N/A" if atype == "str" else None)),
        })

    # Persist optional debug artifacts
    if debug and (debug_csv_path or debug_xlsx_path) and csv_rows:
        import pandas as pd
        dbg_df = pd.DataFrame(csv_rows)
        if debug_csv_path:
            p = Path(debug_csv_path); p.parent.mkdir(parents=True, exist_ok=True)
            dbg_df.to_csv(p, index=False, encoding="utf-8")
            log.info("Wrote debug CSV: %s", p)
        if debug_xlsx_path:
            p = Path(debug_xlsx_path); p.parent.mkdir(parents=True, exist_ok=True)
            dbg_df.to_excel(p, index=False)
            log.info("Wrote debug XLSX: %s", p)

    log.info("QA complete. Total questions: %d", len(results))
    return results

def save_answers_json(answers: List[Dict[str, Any]], out_path: Path):
    out_path.write_text(json.dumps(answers, ensure_ascii=False, indent=2), encoding="utf-8")






from typing import Any, List, Dict
from pathlib import Path
from src.ocr_hybrid import combine_docling_and_openai_for_page  # <-- add import


def _hybrid_ocr_postprocess_pages(
    docling_pages: List[Dict[str, Any]],
    pdf_path: Path,
    *,
    min_chars: int = 200,
    min_confidence: float = 0.60,
    openai_model: str = "gpt-4o-mini",
    dpi: int = 300,
    jpeg_quality: int = 80,
) -> List[Dict[str, Any]]:
    """
    Replace page text with OpenAI OCR if Docling text is too short/low-confidence.
    Expects each page dict to have 'text' and optionally 'confidence' (or 'ocr_confidence').
    """
    for i, p in enumerate(docling_pages):
        dl_text = (p.get("text") or "")
        dl_conf = p.get("confidence", p.get("ocr_confidence"))
        new_text, source = combine_docling_and_openai_for_page(
            str(pdf_path),
            i,
            dl_text,
            dl_conf,
            min_chars=min_chars,
            min_confidence=min_confidence,
            openai_model=openai_model,
            dpi=dpi,
            jpeg_quality=jpeg_quality,
        )
        p["text"] = new_text
        p["ocr_source"] = source
    return docling_pages



import logging
from openai import OpenAI

log = logging.getLogger(__name__)


def safe_chat_create(client: OpenAI, *, model: str, messages, **kwargs):
    """
    Calls chat.completions.create and retries once without 'temperature'
    if the model rejects custom temperature values.
    """
    temp = kwargs.pop("temperature", None)
    try:
        if temp is None:
            return client.chat.completions.create(model=model, messages=messages, **kwargs)
        return client.chat.completions.create(model=model, messages=messages, temperature=temp, **kwargs)
    except BadRequestError as e:
        if "temperature" in str(e).lower():
            log.info("Model '%s' rejected custom temperature. Retrying without 'temperature'.", model)
            return client.chat.completions.create(model=model, messages=messages, **kwargs)
        raise


#########################
#py + ocr


import time
import base64
from pathlib import Path
from typing import Optional

# Optional dependency: PyMuPDF (install: pip install pymupdf)
try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

# OpenAI SDK (already used elsewhere in the project)
try:
    from openai import OpenAI
except Exception:
    OpenAI = None

# Singleton OpenAI client
__ocr_client_singleton: Optional[OpenAI] = None
def _ocr_openai_client() -> OpenAI:
    global __ocr_client_singleton
    if __ocr_client_singleton is None:
        if OpenAI is None:
            raise RuntimeError("openai>=1.40.0 not installed. pip install openai>=1.40.0")
        if not os.getenv("OPENAI_API_KEY"):
            raise RuntimeError("OPENAI_API_KEY is not set for OCR")
        __ocr_client_singleton = OpenAI()
    return __ocr_client_singleton

def _render_page_png(page, dpi: int = 180) -> bytes:
    """
    Render a single PDF page to PNG bytes at given DPI.
    """
    pm = page.get_pixmap(dpi=dpi, alpha=False)
    return pm.tobytes("png")

def _is_lowtext(text: str, min_chars: int = 64) -> bool:
    """
    Heuristic: if extracted text is too short after whitespace compaction, treat as scanned/low text.
    """
    if not text:
        return True
    compact = re.sub(r"\s+", " ", text).strip()
    return len(compact) < min_chars

def _ocr_png_with_openai(png_bytes: bytes, model: str = "gpt-4o-mini", timeout_s: int = 90) -> str:
    """
    OCR a PNG page via OpenAI Vision (chat.completions). Returns plain text.
    Retries transient errors with exponential backoff.
    """
    client = _ocr_openai_client()
    b64 = base64.b64encode(png_bytes).decode("ascii")

    sys_msg = (
        "You are an OCR engine. Transcribe the image to plain text.\n"
        "- Keep the original language; do not translate.\n"
        "- Preserve structure (headings, lists) as plain text.\n"
        "- TABLES: render as GitHub Markdown tables with a header row and an alignment row using '---'.\n"
        "- Use '|' as column separators. Do NOT wrap tables in code fences.\n"
        "- Preserve row/column order; repeat text for merged cells; leave empty cells blank.\n"
        "- No commentary or explanations; output only the transcribed text."
    )
    user_content = [
        {"type": "text", "text": (
            "Transcribe this page. If tables are present, render them as Markdown tables "
            "with a header row and an alignment row. Keep non-table text as plain lines. "
            "Do not add explanations or code fences."
        )},
        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
    ]


    last_err = None
    for attempt in range(4):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": sys_msg},
                    {"role": "user", "content": user_content},
                ],
                temperature=0.0,
                timeout=timeout_s,
            )
            return (resp.choices[0].message.content or "").strip()
        except Exception as e:
            last_err = e
            # backoff: 1s, 2s, 4s
            time.sleep(1.0 * (2 ** attempt))
    # As a last resort, return empty string on failure
    return ""
#

try:
    from tqdm import tqdm as _tqdm
    from tqdm import tqdm as _tqdm_base
    def _tqdm_write(msg: str) -> None:
        try:
            _tqdm_base.write(str(msg))
        except Exception:
            print(str(msg))
except Exception:
    def _tqdm(iterable, **kwargs):
        return iterable
    def _tqdm_write(msg: str) -> None:
        print(str(msg))

# --- replace the page loop in `_parse_pdf_pymupdf_openai(...)` with a tqdm-wrapped loop ---
def _parse_pdf_pymupdf_openai(
    pdf_path: Path,
    out_json_path: Path,
    *,
    ocr_model: str = "gpt-4o-mini",
    image_dpi: int = 180,
    min_chars_for_ocr: int = 64,
    max_pages: Optional[int] = None,
) -> None:
    if fitz is None:
        raise RuntimeError("PyMuPDF (pymupdf) is not installed. pip install pymupdf")

    doc = fitz.open(pdf_path.as_posix())
    total_pages = doc.page_count
    limit = min(total_pages, max_pages) if max_pages else total_pages

    pages: List[Dict[str, Any]] = []
    # Progress per PDF (pages)
    for i in _tqdm(range(limit), desc=f"{pdf_path.name}", unit="pg", leave=False):
        page = doc.load_page(i)
        txt = page.get_text("text") or ""
        used_ocr = False

        if _is_lowtext(txt, min_chars=min_chars_for_ocr):
            try:
                png = _render_page_png(page, dpi=image_dpi)
                ocr_txt = _ocr_png_with_openai(png, model=ocr_model)
                if ocr_txt and len(re.sub(r"\s+", " ", ocr_txt).strip()) >= len(re.sub(r"\s+", " ", txt).strip()):
                    txt = ocr_txt
                    used_ocr = True
            except Exception:
                pass

        pages.append({
            "page_id": i + 1,
            "page_number": i + 1,
            "text": txt or "",
            "ocr_used": used_ocr,
        })

    doc_json = {
        "doc_name": pdf_path.name,
        "source_path": str(pdf_path),
        "num_pages": limit,
        "backend": "pymupdf+openai",
        "ocr_model": ocr_model,
        "content": pages,
    }

    out_json_path.parent.mkdir(parents=True, exist_ok=True)
    with out_json_path.open("w", encoding="utf-8") as f:
        json.dump(doc_json, f, ensure_ascii=False, indent=2)

#  helpers for checkpointed parsing
from typing import Tuple
from pathlib import Path

def _truthy_env(name: str, default: bool = False) -> bool:
    v = os.getenv(name)
    if v is None:
        return default
    v = v.strip().lower()
    return v in ("1", "true", "yes", "y", "on")

def _cached_parse_ok(pdf_path: Path, json_path: Path) -> Tuple[bool, str]:
    if not json_path.exists():
        return False, "no cache"
    try:
        doc = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as e:
        return False, f"bad json ({e})"

    if doc.get("error"):
        return False, "has error"
    if doc.get("interrupted"):
        return False, "interrupted"
    num_pages = int(doc.get("num_pages", 0) or 0)
    content = doc.get("content", [])
    if not isinstance(content, list) or (num_pages and len(content) != num_pages):
        return False, "incomplete content"

    # Skip only if cache is up-to-date vs source PDF
    try:
        if json_path.stat().st_mtime < pdf_path.stat().st_mtime:
            return False, "outdated"
    except Exception:
        # If mtime comparison fails, be conservative and reparse
        return False, "mtime check failed"

    return True, "cached up-to-date"




#  wrap the PDF list loop in `parse_documents_pymupdf_openai(...)` with checkpointing
def parse_documents_pymupdf_openai(
    input_dir: Path,
    output_dir: Path,
    *,
    ocr_model: Optional[str] = None,
    image_dpi: int = 250,
    min_chars_for_ocr: int = 64,
    max_pages: Optional[int] = None,
    skip_existing: bool = None,
    force_reparse: bool = None,
) -> None:
    # env-driven toggles
    if skip_existing is None:
        skip_existing = _truthy_env("RAG_SKIP_EXISTING", True)
    if force_reparse is None:
        force_reparse = _truthy_env("RAG_FORCE_REPARSE", False)

    model = ocr_model or os.getenv("RAG_OCR_MODEL", "gpt-4o-mini")
    try:
        image_dpi = int(os.getenv("RAG_OCR_DPI", str(image_dpi)))
    except Exception:
        pass
    try:
        min_chars_for_ocr = int(os.getenv("RAG_OCR_MIN_CHARS", str(min_chars_for_ocr)))
    except Exception:
        pass
    if os.getenv("RAG_OCR_MAX_PAGES"):
        try:
            max_pages = int(os.getenv("RAG_OCR_MAX_PAGES"))
        except Exception:
            pass

    output_dir.mkdir(parents=True, exist_ok=True)

    pdfs = sorted([p for p in Path(input_dir).rglob("*.pdf") if p.is_file()])
    if not pdfs:
        _tqdm_write(f"[parse_documents_pymupdf_openai] No PDFs found in {input_dir}")
        return

    _tqdm_write(f"[parse_documents_pymupdf_openai] Found {len(pdfs)} PDFs. Parsing into {output_dir} ...")
    skipped = 0
    for pdf in _tqdm(pdfs, desc="Parsing PDFs", unit="pdf"):
        out_json = output_dir / f"{pdf.stem}.json"

        # checkpoint: skip if valid cache exists and not forcing reparse
        if skip_existing and not force_reparse:
            ok, reason = _cached_parse_ok(pdf, out_json)
            if ok:
                skipped += 1
                _tqdm_write(f"  SKIP: {pdf.name} ({reason})")
                continue

        try:
            _parse_pdf_pymupdf_openai(
                pdf,
                out_json,
                ocr_model=model,
                image_dpi=image_dpi,
                min_chars_for_ocr=min_chars_for_ocr,
                max_pages=max_pages,
            )
            _tqdm_write(f"  OK: {pdf.name} -> {out_json.name}")
        except Exception as e:
            stub = {
                "doc_name": pdf.name,
                "source_path": str(pdf),
                "num_pages": 0,
                "backend": "pymupdf+openai",
                "ocr_model": model,
                "content": [],
                "error": str(e),
            }
            with out_json.open("w", encoding="utf-8") as f:
                json.dump(stub, f, ensure_ascii=False, indent=2)
            _tqdm_write(f"  FAIL: {pdf.name} ({e})")

    if skipped:
        _tqdm_write(f"[parse_documents_pymupdf_openai] Skipped {skipped} already parsed PDFs (use RAG_FORCE_REPARSE=1 to reparse)")

#########################
# ==========================
# High-level API
# ==========================

from pathlib import Path
from typing import List
import pandas as pd
import pdfplumber

def _rows_to_df(rows: List[List[str]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    header = rows[0]
    if all(isinstance(c, str) and c.strip() for c in header) and len(rows) > 1:
        return pd.DataFrame(rows[1:], columns=header)
    return pd.DataFrame(rows)

def extract_pdf_tables_to_markdown(input_dir: Path, out_dir: Path) -> List[Path]:
    """
    Extract tables with pdfplumber and write one Markdown file per PDF.
    Returns list of written paths.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    written: List[Path] = []

    pdfs = sorted(p for p in input_dir.rglob("*.pdf") if p.is_file())
    for pdf_path in pdfs:
        md_path = out_dir / f"{pdf_path.stem}_tables.md"
        parts: List[str] = [f"# Tables extracted from {pdf_path.name}\n"]
        count = 0
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for pi, page in enumerate(pdf.pages, start=1):
                    tables = page.extract_tables() or []
                    for ti, rows in enumerate(tables, start=1):
                        df = _rows_to_df(rows)
                        if df.empty:
                            continue
                        count += 1
                        parts.append(f"## Page {pi} — Table {ti}")
                        parts.append(df.to_markdown(index=False))
                        parts.append("")
        except Exception as e:
            parts.append(f"> Extraction failed: {e}")

        if count == 0:
            parts.append("No tables found.")
        md_path.write_text("\n".join(parts), encoding="utf-8")
        written.append(md_path)

    return written



from pathlib import Path

def merge_tables_md_into_parsed(parsed_dir: Path, tables_dir: Path) -> int:
    """
    Append tables Markdown as a synthetic page to each parsed JSON if a matching *_tables.md exists.
    Returns number of JSON files updated.
    """
    updated = 0
    parsed_dir.mkdir(parents=True, exist_ok=True)
    tables_dir.mkdir(parents=True, exist_ok=True)

    for json_path in sorted(parsed_dir.glob("*.json")):
        stem = json_path.stem  # matches '<pdf>.json' <-> '<pdf>_tables.md'
        md_path = tables_dir / f"{stem}_tables.md"
        if not md_path.exists():
            continue

        md_text = md_path.read_text(encoding="utf-8").strip()
        if not md_text or md_text.endswith("No tables found."):
            continue

        doc = json.loads(json_path.read_text(encoding="utf-8"))
        content = doc.get("content", [])
        next_page_num = (content[-1].get("page_number", len(content)) if content else 0) + 1

        content.append({
            "page_id": next_page_num,
            "page_number": next_page_num,
            "text": md_text,
            "ocr_used": False,
            "section": "tables_md"
        })
        doc["content"] = content
        doc["num_pages"] = len(content)

        json_path.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
        updated += 1







######################

from pathlib import Path
from typing import Any, Tuple
import os
import math

try:
    import openpyxl  # pip install openpyxl
except Exception:
    openpyxl = None

# Reuse truthy env helper and cache checker if not already present
def _truthy_env(name: str, default: bool = False) -> bool:
    v = os.getenv(name)
    if v is None:
        return default
    v = v.strip().lower()
    return v in ("1", "true", "yes", "y", "on")

def _cached_parse_ok(src_path: Path, json_path: Path) -> Tuple[bool, str]:
    if not json_path.exists():
        return False, "no cache"
    try:
        doc = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as e:
        return False, f"bad json ({e})"
    if doc.get("error"):
        return False, "has error"
    if doc.get("interrupted"):
        return False, "interrupted"
    num_pages = int(doc.get("num_pages", 0) or 0)
    content = doc.get("content", [])
    if not isinstance(content, list) or (num_pages and len(content) != num_pages):
        return False, "incomplete content"
    try:
        if json_path.stat().st_mtime < src_path.stat().st_mtime:
            return False, "outdated"
    except Exception:
        return False, "mtime check failed"
    return True, "cached up-to-date"

def _cell_to_str(v: Any, max_len: int = 200) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # Render ints without .0 for cleaner text
        if math.isfinite(v) and float(int(v)) == v:
            s = str(int(v))
        else:
            s = f"{v}"
    else:
        s = str(v)
    s = s.replace("\n", " ").strip()
    if len(s) > max_len:
        s = s[: max_len - 1] + "…"
    return s

def _sheet_to_markdown(ws, max_rows: int, max_cols: int) -> str:
    rows_iter = ws.iter_rows(min_row=1, max_row=max_rows or ws.max_row,
                             min_col=1, max_col=max_cols or ws.max_column,
                             values_only=True)
    lines: List[str] = []
    first = True
    for row in rows_iter:
        # Trim trailing empties for compactness
        trimmed = list(row)
        while trimmed and (trimmed[-1] is None or str(trimmed[-1]).strip() == ""):
            trimmed.pop()
        if not trimmed:
            continue
        cells = [_cell_to_str(c) for c in trimmed]
        if first:
            # header + separator to improve retrieval
            lines.append(" | ".join(cells))
            lines.append(" | ".join(["---"] * len(cells)))
            first = False
        else:
            lines.append(" | ".join(cells))
    return "\n".join(lines).strip()

def _parse_xlsx_workbook(
    xlsx_path: Path,
    out_json_path: Path,
    *,
    max_rows_per_sheet: int = None,
    max_cols_per_sheet: int = None,
) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path.as_posix(), data_only=True, read_only=True)
    pages: List[Dict[str, Any]] = []
    try:
        for i, ws in enumerate(wb.worksheets):
            text_md = _sheet_to_markdown(
                ws,
                max_rows=max_rows_per_sheet or int(os.getenv("RAG_XLSX_MAX_ROWS", "1000")),
                max_cols=max_cols_per_sheet or int(os.getenv("RAG_XLSX_MAX_COLS", "100")),
            )
            header = f"# Sheet: {ws.title}".strip()
            txt = f"{header}\n\n{text_md}" if text_md else header
            pages.append({
                "page_id": i + 1,
                "page_number": i + 1,
                "text": txt,
                "ocr_used": False,
                "sheet_name": ws.title,
            })
    finally:
        try:
            wb.close()
        except Exception:
            pass

    doc_json = {
        "doc_name": xlsx_path.name,
        "source_path": str(xlsx_path),
        "num_pages": len(pages),
        "backend": "xlsx+openpyxl",
        "ocr_model": None,
        "content": pages,
    }
    out_json_path.parent.mkdir(parents=True, exist_ok=True)
    with out_json_path.open("w", encoding="utf-8") as f:
        json.dump(doc_json, f, ensure_ascii=False, indent=2)

from typing import Any, Dict, List, Optional, Sequence
from pathlib import Path
import logging

from openai import OpenAI

log = logging.getLogger("RAGPipelineEnsemble")

def _hit_text(h: Dict[str, Any], max_chars: int = 1200) -> str:
    for k in ("chunk_text", "text", "snippet", "content", "body"):
        v = h.get(k)
        if isinstance(v, str) and v.strip():
            s = v.strip().replace("\r", " ")
            return (s[:max_chars] + "…") if len(s) > max_chars else s
    return ""

def _render_first_hit_image_b64(hits: List[Dict[str, Any]], pdf_roots: Sequence[Path], dpi: int = 220) -> Optional[str]:

    print("current_dpi", dpi)
    if not hits or not pdf_roots:
        return None
    # Try the first hit, then the next ones if needed
    for h in hits[:5]:
        page = h.get("page_id") or h.get("page") or h.get("page_number")
        try:
            page = int(page) if page is not None else None
        except Exception:
            page = None
        if not page:
            continue
        pdf = _resolve_pdf_path_from_hit(h, pdf_roots)
        if not pdf:
            continue
        b64 = _render_page_to_png_b64(pdf, page, dpi=dpi)
        if b64:
            return b64
    return None

def _snippets_block(hits: List[Dict[str, Any]], max_ctx_chunks: int = 8, max_snippet_chars: int = 1200) -> str:
    lines = []
    for i, h in enumerate(hits[: max(1, max_ctx_chunks)], start=1):
        doc = h.get("doc_name") or h.get("document_name") or h.get("source") or h.get("doc_stem") or h.get("_index_dir") or ""
        page = h.get("page_id") or h.get("page") or h.get("page_number") or ""
        txt = _hit_text(h, max_snippet_chars)
        lines.append(f"[{i}] doc={doc} page={page}\n{txt}".strip())
    return "\n\n".join(lines) if lines else "(no snippets)"

def _judge_answers(
    client: OpenAI,
    *,
    question: str,
    answer_type: str,
    context_text: str,
    text_answer: Optional[str],
    vision_answer: Optional[str],
    judge_model: str = "gpt-4o-mini",
    max_tokens: int = 250,
) -> Dict[str, Any]:
    """
    Returns: {final_answer: str|None, chosen: 'text'|'vision'|'synth', reason: str}
    """
    sys = (
        "You are an arbiter. Given a question, short context snippets, and two candidate answers "
        "(text-only vs. vision), pick the better one or produce a short synthesized answer.\n"
        "Rules:\n"
        "- Use the context; avoid hallucinations.\n"
        "- If both are wrong/unknown, return 'unknown'.\n"
        "Return strictly JSON: {\"chosen\":\"text|vision|synth\",\"final_answer\":\"...\",\"reason\":\"...\"}."
    )
    user_obj = {
        "question": question,
        "answer_type": answer_type,
        "context_snippets": context_text,
        "candidates": {
            "text": text_answer or "",
            "vision": vision_answer or "",
        },
        "instructions": "Choose the best candidate or synthesize a better one if both are weak."
    }
    resp = client.chat.completions.create(
        model=judge_model,
        messages=[
            {"role": "system", "content": sys},
            {"role": "user", "content": json.dumps(user_obj, ensure_ascii=False)},
        ],
        temperature=0,
        max_tokens=max_tokens,
        response_format={"type": "json_object"},
    )
    raw = (resp.choices[0].message.content or "").strip()
    try:
        obj = json.loads(raw)
    except Exception:
        obj = {}
    chosen = obj.get("chosen") if obj.get("chosen") in ("text", "vision", "synth") else "text"
    final_answer = obj.get("final_answer")
    if final_answer is not None and not isinstance(final_answer, str):
        final_answer = str(final_answer)
    reason = obj.get("reason") or ""
    return {"final_answer": final_answer, "chosen": chosen, "reason": reason, "raw": raw}



############################################
# candidate a

from typing import Any, Dict, List, Optional, Tuple
import re
import logging

from openai import OpenAI
try:
    from openai import BadRequestError  # available in your file already
except Exception:
    BadRequestError = Exception  # fallback to generic

log = logging.getLogger(__name__)

# If present elsewhere in the file, this will reuse it; otherwise define a minimal set.
try:
    UNSUPP_TEMP_MODELS
except NameError:
    UNSUPP_TEMP_MODELS = {"gpt-5", "gpt-5-mini", "gpt-5.1", "gpt-5.1-mini"}

def _strip_code_fences(s: str) -> str:
    s = s.strip()
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?\s*", "", s)
        s = re.sub(r"\s*```$", "", s)
    return s.strip()

def _extract_json_obj(text: str) -> Optional[Dict[str, Any]]:
    # Try to find the first JSON object in the text.
    text = text.strip()
    # Fast path: direct JSON
    try:
        obj = json.loads(_strip_code_fences(text))
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass
    # Fallback: locate {...} by bracket counting
    start = text.find("{")
    while start != -1:
        depth = 0
        for i in range(start, len(text)):
            ch = text[i]
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    snippet = text[start:i+1]
                    try:
                        obj = json.loads(snippet)
                        if isinstance(obj, dict):
                            return obj
                    except Exception:
                        break
        start = text.find("{", start + 1)
    return None



def _llm_answer_with_support_candidate_a(
    client: OpenAI,
    model: str,
    question: str,
    answer_type: str,
    hits: List[Dict[str, Any]],
    max_ctx_chunks: int = 8,
    max_snippet_chars: int = 1800,
    image_b64: Optional[str] = None,
) -> Tuple[Optional[str], Optional[int], Optional[str]]:
    ctx_parts: List[str] = []
    for i, h in enumerate(hits[: max_ctx_chunks], start=1):
        src = h.get("source") or h.get("_index_dir") or h.get("doc_name") or "unknown"
        page = h.get("page_id", "?")
        text = (h.get("text") or "").strip().replace("\r", " ")
        if len(text) > max_snippet_chars:
            text = text[:max_snippet_chars] + "…"
        ctx_parts.append(f"[{i}] {src} (page {page})\n{text}")
    context_block = "\n\n---\n\n".join(ctx_parts) if ctx_parts else "(no context)"

    at = (answer_type or "str").strip().lower()
    if at not in ("str", "int", "float"):
        at = "str"

    sys_msg = (
        "You are a precise QA assistant for financial and corporate documents.\n"
        "Return EXACTLY one JSON object: {\"answer\": <value>, \"support\": <passage_index>}.\n"
        "Rules (apply in order):\n"
        "1. ALWAYS output a non-null answer (best plausible hypothesis / 'предположение') even if evidence is weak.\n"
        "2. Never output null or N/A.\n"
        "3. support: single integer 1..N of the passage that MOST directly supports the answer; if truly no passage applies, use null.\n"
        "4. answer_type=int  -> digits only (no spaces, no separators, sign only if negative).\n"
        "5. answer_type=float -> '.' decimal separator, no thousands separators; trim trailing zeros; no scientific notation.\n"
        "6. answer_type=str -> concise string (no enclosing quotes); for Russian answers use nominative case (именительный падеж).\n"
        "7. ROLE / TITLE QUERIES (question starts with / contains: 'Кто', 'Какую должность', 'Какова должность', 'Должность', 'Какая должность'): OUTPUT ONLY the pure job title words. STRICTLY EXCLUDE:\n"
        "   - Company legal forms and names (АО, ТОО, LLP, AO, OOO, АО «...», ТОО \"...\").\n"
        "   - Trailing company specification after the title (e.g., remove 'АО «ЛОТТЕ Рахат»').\n"
        "   - Department/company repetition unless it is an intrinsic part of the title (e.g., keep 'директор по производству', drop company name).\n"
        "   Only include company name if the question explicitly asks for the company or disambiguation is impossible without it.\n"
        "   Example:\n"
        "     Q: Какую должность занимает Им Вонхёк в АО \"ЛОТТЕ Рахат\"?\n"
        "     Passage: 'Заместитель председателя правления по производству АО «ЛОТТЕ Рахат»'\n"
        "     Answer: 'Заместитель председателя правления по производству'\n"
        "8. Do NOT repeat the full company name if the question already specifies it; output only the requested value.\n"
        "9. Preserve original internal punctuation of the chosen value; do not delete, insert, or alter punctuation characters within it.\n"
        "10. Proper names (organizations, programs, documents, geographic, personal names, etc.) must be reproduced EXACTLY (casing, hyphens, quotes, apostrophes, dots, symbols) without normalization or reordering.\n"
        "11. If multiple numeric candidates exist, choose the one whose immediate context best matches the question semantics.\n"
        "12. If you must infer, pick the most reasonable value derivable from provided passages (state it directly; do NOT label it inside the value).\n"
        "13. Use ONLY the provided passages (and page image if present); no outside knowledge.\n"
        "14. Do NOT invent units; output raw value only.\n"
        "15. Output ONLY the JSON object; no markdown, no extra keys, no commentary.\n"
        "16. Trim leading/trailing spaces in the answer.\n"
        "Produce JSON now."
    )
    base_user_text = (
        f"question: {question}\n"
        f"answer_type: {at}\n\nPASSAGES:\n{context_block}\n\nOutput JSON only:"
    )

    # Build multi-modal or text-only content
    if image_b64:
        user_content = [
            {"type": "text", "text": base_user_text},
            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}},
        ]
    else:
        user_content = base_user_text  # plain string

    params: Dict[str, Any] = {
        "model": model,
        "messages": [
            {"role": "system", "content": sys_msg},
            {"role": "user", "content": user_content},
        ],

        "seed" : 2025+1,
    }
    if model not in UNSUPP_TEMP_MODELS:
        params["temperature"] = 0

    try:
        resp = client.chat.completions.create(**params)
    except BadRequestError as e:
        if "temperature" in params:
            params.pop("temperature", None)
        resp = client.chat.completions.create(**params)
    except Exception:
        if "temperature" in params:
            params.pop("temperature", None)
        try:
            resp = client.chat.completions.create(**params)
        except Exception:
            return None, None, None

    raw = (resp.choices[0].message.content or "").strip()
    obj = _extract_json_obj(raw) or {}
    ans = obj.get("answer", None)
    ans_str = None if ans is None else str(ans).strip()

    sup = obj.get("support", None)
    try:
        sup_idx = int(sup) if sup is not None else None
    except Exception:
        sup_idx = None

    max_allowed = min(len(hits), max_ctx_chunks)
    if sup_idx is not None and (sup_idx < 1 or sup_idx > max_allowed):
        sup_idx = None

    return ans_str, sup_idx, raw

# Simplified: only primary (text) model used; vision + judge disabled.
def ensemble_answer(
    client: OpenAI,
    *,
    question: str,
    answer_type: str,
    hits: List[Dict[str, Any]],
    pdf_roots: Optional[Sequence[Path]] = None,
    max_ctx_chunks: int = 8,
    max_snippet_chars: int = 1200,
    model_text: str = "gpt-5",
    model_vision: str = "gpt-4o",      # retained for signature compatibility (unused)
    judge_model: str = "gpt-4o-mini",  # retained (unused)
    text_with_image: bool = True,
) -> Dict[str, Any]:
    # Optionally render one page image (only if we still want to feed it to the text model)
    image_b64 = None
    if text_with_image and pdf_roots:
        image_b64 = _render_first_hit_image_b64(hits, pdf_roots or [], dpi=300)

    # Single (text) candidate answer
    ans_text, sup_text, raw_text = _llm_answer_with_support_candidate_a(
        client=client,
        model=model_text,
        question=question,
        answer_type=answer_type,
        hits=hits,
        max_ctx_chunks=max_ctx_chunks,
        max_snippet_chars=max_snippet_chars,
        image_b64=image_b64 if text_with_image else None,
    )

    coerced = _coerce_answer(ans_text, answer_type)

    # Support fallback if model did not provide one
    sup_1b: Optional[int] = sup_text
    if sup_1b is None:
        sup_1b = _guess_support_by_value(hits, coerced) or (1 if hits else None)

    return {
        "final_answer": coerced,
        "chosen": "text",
        "reason": "",
        "support_1b": sup_1b,
        "raw": {
            "text": raw_text,
            "vision": None,  # kept for compatibility
            "judge": None,
        },
        "candidates": {
            "text": ans_text,
            "vision": None,
        },
        "image_used": bool(image_b64),
        "text_with_image": text_with_image,
    }



from typing import Any, Dict, List, Optional, Sequence
from pathlib import Path
import json

from openai import OpenAI

def answer_questions_from_xlsx_configurable_models(
    df,
    retriever,
    parsed_dir: Path,
    *,
    # ensemble models (exposed as parameters)
    model_text: str = "gpt-5",          # chunks only
    model_vision: str = "gpt-4o",       # chunks + image
    judge_model: str = "gpt-4o-mini",   # arbiter
    # retrieval / context parameters
    top_k_per_doc: int = 20,
    max_ctx_chunks: int = 8,
    max_snippet_chars: int = 1200,
    pdf_roots: Optional[Sequence[Path]] = None,
    # debug / logging
    debug: bool = True,
    debug_snippet_chars: int = 220,
    debug_print_top_k: int = 10,
    # client
    client: Optional[OpenAI] = None,
) -> List[Dict[str, Any]]:
    """
    Runs RAG ensemble answering with fully configurable models:
    - model_text: LLM for text-only answers (uses only chunks)
    - model_vision: vision-capable LLM for chunks + page image
    - judge_model: small LLM to pick/synthesize the final answer
    Returns a list of {question_id, relevant_chunks:[{document_name,page_number}], answer}.
    """
    from tqdm import tqdm

    # Lazy imports for helpers expected in this module
    # - ensemble_answer
    # - _guess_support_by_value (used indirectly)
    # - Any rendering helpers used by ensemble_answer are reused
    if client is None:
        client = OpenAI()

    # df = pd.read_excel(xlsx_path)
    cols = {c.lower(): c for c in df.columns}
    for required in ("id", "full_question", "answer_type"):
        if required not in cols:
            raise ValueError(f"XLSX must contain column '{required}'")

    results: List[Dict[str, Any]] = []

    for _, row in tqdm(df.iterrows(), total=len(df), desc="Answers(ensemble-config)"):
        qid = int(row[cols["id"]])
        qtext = str(row[cols["full_question"]]).strip()
        atype = str(row[cols["answer_type"]]).strip().lower()

        # 1) retrieve
        hits = retriever.search_all_docs(qtext, top_k=top_k_per_doc) or []

        # 2) ensemble with configurable models
        ens = ensemble_answer(
            client,
            question=qtext,
            answer_type=atype,
            hits=hits,
            pdf_roots=pdf_roots,
            max_ctx_chunks=max_ctx_chunks,
            max_snippet_chars=max_snippet_chars,
            model_text=model_text,         # <- chunks only
            model_vision=model_vision,     # <- chunks + image
            judge_model=judge_model,       # <- judge/arbiter
            text_with_image=True
        )

        # 3) choose one supporting chunk (1-based)
        rels: List[Dict[str, Any]] = []
        sup = ens.get("support_1b")
        if hits and isinstance(sup, int) and 1 <= sup <= len(hits):
            chosen_hit = hits[sup - 1]
            src_hint = chosen_hit.get("_index_dir") or chosen_hit.get("doc_name")
            from pathlib import Path as _P
            try:
                name = None
                p = _P(str(src_hint)) if src_hint else None
                if p and p.is_dir() and (p / "info.json").exists():
                    info = json.loads((p / "info.json").read_text(encoding="utf-8"))
                    name = info.get("source_path") or info.get("pdf_path") or info.get("doc_stem")
                if not name:
                    name = chosen_hit.get("document_name") or chosen_hit.get("doc_name") or (p.name if p else "unknown")
                page = int(chosen_hit.get("page_id", chosen_hit.get("page_number", -1)))
            except Exception:
                name, page = "unknown", int(chosen_hit.get("page_id", -1))
            rels.append({"document_name": Path(str(name)).name, "page_number": page})

        # 4) result item (keeps the same output schema)
        results.append({
            "question_id": qid,
            "relevant_chunks": rels,  # exactly one
            "answer": ens.get("final_answer") if ens.get("final_answer") is not None else ("N/A" if atype == "str" else None),
        })

        # 5) debug
        if debug:
            print("\n" + "=" * 80)
            print(f"QID {qid}")
            print(f"Q: {qtext}")
            print(f"Type: {atype}")
            print(f"Models -> text={model_text} vision={model_vision} judge={judge_model}")
            print("-" * 80)
            print("Top hits:")
            for i, h in enumerate(hits[:debug_print_top_k], start=1):
                score = float(h.get("final_score", h.get("score", 0.0)))
                src = h.get("_index_dir") or h.get("doc_name") or "unknown"
                page = h.get("page_id", "?")
                text = (h.get("text") or "").replace("\n", " ")
                snippet = text[:debug_snippet_chars] + ("…" if len(text) > debug_snippet_chars else "")
                print(f"[{i}] {src} p.{page} score={score:.4f} | {snippet}")
            print("-" * 80)
            cands = ens.get("candidates", {})
            print(f"Text ans:   {cands.get('text')}")
            print(f"Vision ans: {cands.get('vision')}")
            print(f"Judge -> chosen={ens.get('chosen')} reason={ens.get('reason')}")
            if rels:
                outs = ", ".join([f"{c.get('document_name')}:{c.get('page_number')}" for c in rels])
                print(f"Used chunk: {outs}")
            else:
                print("Used chunk: (none)")

    return results###################


#  XLSX parsing support (openpyxl)

from pathlib import Path
from typing import Any, List, Dict, Optional

try:
    import openpyxl  # pip install openpyxl
except Exception:
    openpyxl = None


def _cell_to_str(v: Any, max_len: int = 200) -> str:
    """
    Convert any Excel cell value to a compact string suitable for Markdown table output.
    Trims long values and normalizes whitespace.
    """
    if v is None:
        s = ""
    elif isinstance(v, bool):
        s = "TRUE" if v else "FALSE"
    else:
        s = str(v)
    s = " ".join(s.split())
    if len(s) > max_len:
        s = s[: max_len - 1] + "…"
    return s


def _sheet_to_markdown(ws, max_rows: int, max_cols: int) -> str:
    """
    Render a worksheet to a simple Markdown-like table.
    Limits rows/cols to avoid giant prompts.
    """
    rows_limit = min(ws.max_row or 0, max_rows)
    cols_limit = min(ws.max_column or 0, max_cols)

    lines: List[str] = []
    lines.append(f"# Sheet: {ws.title}")
    if rows_limit == 0 or cols_limit == 0:
        lines.append("(empty sheet)")
        return "\n".join(lines)

    # Build table header as A,B,C,... for readability
    col_headers = [openpyxl.utils.get_column_letter(c) for c in range(1, cols_limit + 1)]
    lines.append("| " + " | ".join(col_headers) + " |")
    lines.append("| " + " | ".join(["---"] * cols_limit) + " |")

    for row in ws.iter_rows(min_row=1, max_row=rows_limit, min_col=1, max_col=cols_limit, values_only=True):
        vals = [_cell_to_str(v) for v in row]
        lines.append("| " + " | ".join(vals) + " |")

    if ws.max_row > rows_limit or ws.max_column > cols_limit:
        lines.append("")
        lines.append(f"(truncated to {rows_limit} rows × {cols_limit} cols)")

    return "\n".join(lines)


def _parse_xlsx_workbook(xlsx_path: Path, out_json_path: Path, *, max_rows: int = 300, max_cols: int = 40) -> None:
    """
    Parse a single .xlsx into a V2-like JSON:
      - One worksheet -> one 'page' with Markdown table text.
      - Fields: doc_name, source_path, backend='openpyxl', content=[{page_id,page_number,text,sheet_title}]
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path.as_posix(), data_only=True, read_only=True)
    pages: List[Dict[str, Any]] = []
    page_id = 1
    for ws in wb.worksheets:
        text_md = _sheet_to_markdown(ws, max_rows=max_rows, max_cols=max_cols)
        pages.append({
            "page_id": page_id,
            "page_number": page_id,
            "text": text_md,
            "sheet_title": ws.title,
        })
        page_id += 1

    doc_json = {
        "doc_name": xlsx_path.name,
        "source_path": str(xlsx_path),
        "num_pages": len(pages),
        "backend": "openpyxl",
        "ocr_model": None,
        "content": pages,
    }

    out_json_path.parent.mkdir(parents=True, exist_ok=True)
    out_json_path.write_text(json.dumps(doc_json, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_workbooks_openpyxl(input_dir: Path, output_dir: Path, *, max_rows: int = 300, max_cols: int = 40) -> int:
    """
    Batch-parse all .xlsx files under input_dir into output_dir.
    Returns the number of JSON files written.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Install with: pip install openpyxl")

    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_files = sorted(p for p in Path(input_dir).rglob("*.xlsx") if p.is_file())
    if not xlsx_files:
        logging.getLogger("RAGPipeline").info("No .xlsx files found in %s", input_dir)
        return 0

    ok = 0
    for xlsx in xlsx_files:
        # Skip temp/hidden files
        name = xlsx.name
        if name.startswith("~$") or name.startswith("."):
            continue

        out_json = output_dir / f"{xlsx.stem}.json"
        try:
            _parse_xlsx_workbook(xlsx, out_json, max_rows=max_rows, max_cols=max_cols)
            ok += 1
            logging.getLogger("RAGPipeline").info("Parsed workbook: %s -> %s", xlsx.name, out_json.name)
        except Exception as e:
            # Write a stub to keep pipeline robust
            stub = {
                "doc_name": xlsx.name,
                "source_path": str(xlsx),
                "num_pages": 0,
                "backend": "openpyxl",
                "ocr_model": None,
                "content": [],
                "error": str(e),
            }
            out_json.write_text(json.dumps(stub, ensure_ascii=False, indent=2), encoding="utf-8")
            logging.getLogger("RAGPipeline").warning("Failed workbook: %s (%s)", xlsx.name, e)

    return ok


def build_corpus_and_indices(parse_cfg: ParseConfig, indexer_cfg: IndexerConfig) -> List[Path]:
    parser = CorpusParser(parse_cfg)
    parser.run()
    indexer = RAGIndexer(indexer_cfg)
    return indexer.build_all()


__all__ = [
    "ParseConfig", "IndexerConfig", "ChunkingConfig", "HybridConfig",
    "RAGIndexer", "RAGRetriever", "RetrieverConfig", "RerankConfig",
    "QuestionPrepConfig",
    "answer_questions_from_xlsx", "save_answers_json",
    "build_corpus_and_indices",  'extract_pdf_tables_to_markdown',
    'merge_tables_md_into_parsed', 'parse_workbooks_openpyxl', "parse_workbooks_openpyxl",
    'parse_documents_pymupdf_openai'
]
